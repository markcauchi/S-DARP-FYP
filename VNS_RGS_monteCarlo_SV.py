import numpy as np
import random
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
from matplotlib.lines import Line2D
import itertools
import copy
import datetime
import time

class Request:
    def __init__(self, requestNo, pickupNode, deliveryNode, pickupTimeWindowE, pickupTimeWindowL, deliveryTimeWindowE,
                 deliveryTimeWindowL, directDistance, reward, load):
        self.requestNo = requestNo
        self.pickupNode = pickupNode
        self.deliveryNode = deliveryNode
        self.pickupTimeWindowE = pickupTimeWindowE
        self.pickupTimeWindowL = pickupTimeWindowL
        self.deliveryTimeWindowE = deliveryTimeWindowE
        self.deliveryTimeWindowL = deliveryTimeWindowL
        self.directDistance = directDistance
        self.reward = reward
        self.load = load

def newSolutionPlot(currentRoute, currentRouteRequest, currentRouteTime, currentRouteLoad, currentReward, currentCost,
                    dataOfImages):
    fig, ax = plt.subplots()
    imgplot = plt.imshow(img)
    plt.axis('off')

    # profits with the current solution
    currentProfit = round(currentReward - currentCost, 2)
    currentReward = round(currentReward, 2)
    currentCost = round(currentCost, 2)
    title = "Total Profit is: " + str(currentProfit) + "\nTotal Reward is: " + str(currentReward) + \
            "\nTotal Cost is: " + str(currentCost)
    plt.title(title)

    # legend
    legend_elements = [Line2D([0], [0], marker='o', color='g', label='Unserved Requests',
                              linestyle='', markerfacecolor='g', markersize=6),
                       Line2D([0], [0], marker='o', color='b', label='Served Pickup Nodes',
                              linestyle='', markerfacecolor='b', markersize=6),
                       Line2D([0], [0], marker='o', color='r', label='Served Delivery Nodes',
                              linestyle='', markerfacecolor='r', markersize=6),
                       Line2D([0], [0], marker='o', color='indigo', label='Driver Start/End Nodes',
                              linestyle='', markerfacecolor='indigo', markersize=6),
                       Line2D([0], [0], color='c', label='Number on line = Load')]

    # show all available nodes which requests may start/end at
    for i in range(noOfNodes):
        plt.scatter(imgNums[i][0], imgNums[i][1], c='g', s=20)

    # plotting of routes
    for DN in range(len(currentRoute)):  # for all drivers
        for i in range(len(currentRoute[DN]) - 1):
            # plot a line to represent each route
            plt.plot([imgNums[currentRoute[DN][i] - 1][0], imgNums[currentRoute[DN][i + 1] - 1][0]],
                     [imgNums[currentRoute[DN][i] - 1][1], imgNums[currentRoute[DN][i + 1] - 1][1]], '-c')

            # annotate the load on each edge
            label = currentRouteLoad[DN][i]
            plt.annotate(label, ((imgNums[currentRoute[DN][i] - 1][0] + imgNums[currentRoute[DN][i+1] - 1][0])/2,
                                 (imgNums[currentRoute[DN][i] - 1][1] + imgNums[currentRoute[DN][i+1] - 1][1])/2),
                         textcoords="offset points", xytext=(0, -8), ha='center', size=8)

    # colour code the drivers' depots, the pickup points and the delivery points
    for DN in range(len(currentRoute)):
        pickupNodes = []
        for i in range(len(currentRoute[DN])):
            if currentRouteRequest[DN][i] == driverNo[DN]:  # depots - indigo
                plt.scatter(imgNums[currentRoute[DN][i] - 1][0], imgNums[currentRoute[DN][i] - 1][1], c='indigo', s=60)
            elif currentRouteRequest[DN][i] not in pickupNodes:  # delivery nodes - blue
                pickupNodes.append(currentRouteRequest[DN][i])
                plt.scatter(imgNums[currentRoute[DN][i] - 1][0], imgNums[currentRoute[DN][i] - 1][1], c='b', s=40)
            else:  # pickup nodes - red
                plt.scatter(imgNums[currentRoute[DN][i] - 1][0], imgNums[currentRoute[DN][i] - 1][1], c='r', s=40)
            # plt.scatter(imgNums[i][0], imgNums[i][1], c='g', s=40)

    # add the legend to the plot
    ax.legend(handles=legend_elements, loc=1, fontsize=7.2)
    plt.savefig('routeImage.png', bbox_inches='tight')
    img1 = mpimg.imread('routeImage.png')
    dataOfImages.append(img1)
    plt.show()
    return dataOfImages


def obtainScales (pltMalta):
    # taking the size of the image
    bottom, top = pltMalta.ylim()
    left, right = pltMalta.xlim()

    bottom += 0.5
    right += 0.5

    # x, y format
    imageLimits = np.array([right, bottom])  # scale of image coordinates

    # fixed coordinates of the image shown
    geographicBottom = 35.848350
    geographicTop = 35.932180
    geographicLeft = 14.399522
    geographicRight = 14.539597

    # x, y format
    # scale of geographic coordinates
    geographicLimits = np.array([geographicRight - geographicLeft, geographicBottom - geographicTop])
    return imageLimits, geographicLimits, geographicLeft, geographicTop


def normaliseGeo (geoLocation, geoL, geoT):
    # used for finding a linear relationship between geographical coordinates and image pixels
    location = np.array([geoLocation[0] - geoL, geoLocation[1] - geoT])
    return location


def convertGeoToImg (geoLocation, geoScale, imgScale):
    # find the image pixels of where the geographical coordinates are pointing to
    imgLocation = np.multiply(np.divide(geoLocation, geoScale), imgScale)
    return imgLocation


def convertImgToGeo (imgLocation, geoScale, imgScale, geoT, geoL):
    # find the geographical coordinates of an image pixel to find its respective point on map
    geoLocation = np.multiply(np.divide(imgLocation, imgScale), geoScale) + np.array([geoL, geoT])
    return geoLocation


def deleteRequestFromRoute(xPrime, xPrimeRequest, requestToDelete):
    matches = [k for k, x in enumerate(xPrimeRequest) if x == requestToDelete]
    del xPrime[matches[1]]
    del xPrime[matches[0]]
    del xPrimeRequest[matches[1]]
    del xPrimeRequest[matches[0]]
    return xPrime, xPrimeRequest


def findBestDeletion(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeReward, xPrimeCost, tMatrix, allRequests):
    bestProfit = sum(xPrimeReward) - sum(xPrimeCost)
    bestRoute = copy.deepcopy(xPrime)
    bestRouteRequest = copy.deepcopy(xPrimeRequest)

    for DN in range(len(xPrime)):  # for each route
        for i in range(len(xPrimeServed[DN])):  # for each request served inside route
            if xPrimeServed[DN][i] != driverNo[DN]:  # driver's request must not be removed
                # find where the nodes of the request considered are located in the route
                consideredRequest = xPrimeServed[DN][i]
                x2Prime = copy.deepcopy(xPrime)
                x2PrimeRequest = copy.deepcopy(xPrimeRequest)
                x2PrimeReward = xPrimeReward[:]
                x2PrimeCost = xPrimeCost[:]

                # delete the nodes serving that request
                [x2Prime[DN], x2PrimeRequest[DN]] = deleteRequestFromRoute(x2Prime[DN], x2PrimeRequest[DN],
                                                                           consideredRequest)

                # update the information of the new route
                [x2Prime[DN], x2PrimeRequest[DN], tempRouteLoad, tempRouteTimeWasted, tempRouteTime,
                 x2PrimeReward[DN], x2PrimeCost[DN], tempPartiallyServedRequests, tempServedRequests, isItPossible] = \
                    updateRouteInfo(x2Prime[DN], x2PrimeRequest[DN], xPrimeTime[DN], tMatrix, allRequests)

                if sum(x2PrimeReward) - sum(x2PrimeCost) > bestProfit and isItPossible:  # if an improvement is found
                    # replace the best
                    bestProfit = sum(x2PrimeReward) - sum(x2PrimeCost)
                    bestRoute = copy.deepcopy(x2Prime)
                    bestRouteRequest = copy.deepcopy(x2PrimeRequest)

    return bestRoute, bestRouteRequest, bestProfit

def findBestIntraSwap(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeReward, xPrimeCost,
                      tMatrix, allRequests, numOfRequests):
    bestRoute = copy.deepcopy(xPrime)
    bestRouteRequest = copy.deepcopy(xPrimeRequest)

    for DN in range(len(xPrime)):  # for each route
        bestProfitI = xPrimeReward[DN] - xPrimeCost[DN]
        bestRouteI = copy.deepcopy(bestRoute[DN])
        bestRouteRequestI = copy.deepcopy(bestRouteRequest[DN])

        for i in range(len(xPrimeServed[DN])):  # for each request served
            if xPrimeServed[DN][i] != driverNo[DN]:  # driver start and end nodes must never be changed
                # find where the nodes of the considered request are located in the route
                consideredRequest = xPrimeServed[DN][i]
                x2Prime = copy.deepcopy(bestRouteI)
                x2PrimeRequest = copy.deepcopy(bestRouteRequestI)

                # remove the request from the route
                [x2Prime, x2PrimeRequest] = deleteRequestFromRoute(x2Prime, x2PrimeRequest, consideredRequest)

                # create a temporary variable that allows only the removed request to be unserved
                tempServedRequests = list(range(1, numOfRequests + 1))
                del tempServedRequests[xPrimeServed[DN][i] - 1]

                totalReward = 0
                totalCost = float("inf")

                # find best insertion will only try to add the removed request
                [x2Prime, x2PrimeRequest, x2PrimeCost, x2PrimeReward] =\
                    findBestInsertionSV(x2Prime, x2PrimeRequest, xPrimeTime[DN], totalReward, totalCost,
                                        tempServedRequests, tMatrix, allRequests, numOfRequests)
                x2PrimeProfitI = x2PrimeReward - x2PrimeCost
                if x2PrimeProfitI > bestProfitI:  # if we found an improvement
                    # replace the best
                    bestProfitI = x2PrimeProfitI
                    bestRoute[DN] = copy.deepcopy(x2Prime)
                    bestRouteRequest[DN] = copy.deepcopy(x2PrimeRequest)

    return bestRoute, bestRouteRequest


def findBestMove(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeReward, xPrimeCost,
                 xPrimeRewardTotal, xPrimeCostTotal, tMatrix, allRequests, numOfRequests):
    bestProfit = xPrimeRewardTotal - xPrimeCostTotal
    bestRoute = copy.deepcopy(xPrime)
    bestRouteRequest = copy.deepcopy(xPrimeRequest)

    for DN in range(len(xPrime)):  # for each route
        for i in range(len(xPrimeServed[DN])):  # for each request served in that route
            if xPrimeServed[DN][i] != driverNo[DN]:  # driver's request must not be moved
                # find where the request's nodes are located
                consideredRequest = xPrimeServed[DN][i]
                x2Prime = copy.deepcopy(xPrime)
                x2PrimeRequest = copy.deepcopy(xPrimeRequest)
                x2PrimeCost = copy.deepcopy(xPrimeCost)
                x2PrimeReward = copy.deepcopy(xPrimeReward)

                # remove the request from the route
                [x2Prime[DN], x2PrimeRequest[DN]] = deleteRequestFromRoute(x2Prime[DN], x2PrimeRequest[DN],
                                                                           consideredRequest)

                # create a temporary variable that indicates that the removed request is the only unserved request
                tempServedRequests = list(range(1, numOfRequests + 1))
                del tempServedRequests[xPrimeServed[DN][i] - 1]

                for j in range(len(xPrime)):  # for each route
                    if DN != j:  # do not consider the current route
                        x2PrimeTemp = copy.deepcopy(x2Prime)
                        x2PrimeRequestTemp = copy.deepcopy(x2PrimeRequest)
                        x2PrimeCostTemp = x2PrimeCost[:]
                        x2PrimeRewardTemp = xPrimeReward[:]
                        x2PrimeCheckRoute = x2Prime[j][:]
                        dummyReward = 0
                        dummyCost = float("inf")
                        # try to add the removed request inside another route
                        [x2PrimeTemp[j], x2PrimeRequestTemp[j], x2PrimeCostTemp[j], x2PrimeRewardTemp[j]] =\
                            findBestInsertionSV(x2PrimeTemp[j], x2PrimeRequestTemp[j], xPrimeTime[j], dummyReward,
                                                dummyCost, tempServedRequests, tMatrix, allRequests, numOfRequests)

                        # if it managed to add the request without violating any constraints
                        if x2PrimeCheckRoute != x2PrimeTemp[j]:
                            # update the information on all routes
                            [x2PrimeTemp, x2PrimeRequestTemp, x2PrimeLoadTemp, x2PrimeTimeWastedTemp,
                             x2PrimeTimeTemp, x2PrimeRewardTemp, x2PrimeCostTemp, x2PrimeFinalRewardTemp,
                             x2PrimeFinalCostTemp, x2PrimePartiallyServedTemp, x2PrimeServedTemp,
                             tempPartiallyServedRequestsAllVehicles, tempServedRequestsAll, isItPossible] = \
                                updateAllRoutes(x2PrimeTemp, x2PrimeRequestTemp, xPrimeTime, tMatrix, allRequests)

                            # check if profits have increased
                            x2PrimeProfitTemp = sum(x2PrimeRewardTemp) - sum(x2PrimeCostTemp)
                            if x2PrimeProfitTemp > bestProfit and isItPossible:  # if we found an improvement
                                # replace the best
                                bestProfit = x2PrimeProfitTemp
                                bestRoute = copy.deepcopy(x2PrimeTemp)
                                bestRouteRequest = copy.deepcopy(x2PrimeRequestTemp)

    return bestRoute, bestRouteRequest, bestProfit


def findBestRouteSwap(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeReward, xPrimeCost,
                      xPrimeRewardTotal, xPrimeCostTotal, tMatrix, allRequests, numOfRequests):
    bestProfit = xPrimeRewardTotal - xPrimeCostTotal
    bestRoute = copy.deepcopy(xPrime)
    bestRouteRequest = copy.deepcopy(xPrimeRequest)

    for DN in range(len(xPrime)):  # for each route
        for i in range(len(xPrimeServed[DN])):  # for each request served in that route
            if xPrimeServed[DN][i] != driverNo[DN]:  # driver's request must not be replaced
                # find where the considered request is located in current route
                consideredRequest = xPrimeServed[DN][i]
                x2Prime = copy.deepcopy(xPrime)
                x2PrimeRequest = copy.deepcopy(xPrimeRequest)
                x2PrimeCost = copy.deepcopy(xPrimeCost)
                x2PrimeReward = copy.deepcopy(xPrimeReward)

                # remove request from the route
                [x2Prime[DN], x2PrimeRequest[DN]] = deleteRequestFromRoute(x2Prime[DN], x2PrimeRequest[DN],
                                                                           consideredRequest)

                # create a temporary variable that indicates that the removed request is the only unserved one
                tempServedRequests = list(range(1, numOfRequests + 1))
                del tempServedRequests[xPrimeServed[DN][i] - 1]

                for j in range(DN + 1, len(xPrime)):  # for the rest of the drivers/routes
                    for m in range(len(xPrimeServed[j])):  # for each request served in the next route
                        if xPrimeServed[j][m] != driverNo[j]:  # driver's request must never be replaced
                            # find where the considered request is located
                            consideredRequest2 = xPrimeServed[j][m]
                            x2PrimeTemp = copy.deepcopy(x2Prime)
                            x2PrimeRequestTemp = copy.deepcopy(x2PrimeRequest)
                            x2PrimeCostTemp = x2PrimeCost[:]
                            x2PrimeRewardTemp = xPrimeReward[:]
                            # remove the considered request
                            [x2PrimeTemp[j], x2PrimeRequestTemp[j]] = \
                                deleteRequestFromRoute(x2PrimeTemp[j], x2PrimeRequestTemp[j], consideredRequest2)

                            x2PrimeCheckRoute = x2PrimeTemp[j][:]
                            dummyReward = 0
                            dummyCost = float("inf")

                            # try to insert the removed request from the previous route into this one
                            [x2PrimeTemp[j], x2PrimeRequestTemp[j], x2PrimeCostTemp[j], x2PrimeRewardTemp[j]] =\
                                findBestInsertionSV(x2PrimeTemp[j], x2PrimeRequestTemp[j], xPrimeTime[j], dummyReward,
                                                    dummyCost, tempServedRequests, tMatrix, allRequests, numOfRequests)

                            # if a feasible insertion was found
                            if x2PrimeCheckRoute != x2PrimeTemp[j]:
                                x2PrimeCheckRoute2 = x2PrimeTemp[DN][:]
                                tempServedRequests2 = list(range(1, numOfRequests + 1))
                                del tempServedRequests2[xPrimeServed[j][m] - 1]

                                # try to insert the removed request from this route into the previous one
                                [x2PrimeTemp[DN], x2PrimeRequestTemp[DN], x2PrimeCostTemp[DN],
                                 x2PrimeRewardTemp[DN]] = findBestInsertionSV(x2PrimeTemp[DN], x2PrimeRequestTemp[DN],
                                                                              xPrimeTime[DN], dummyReward, dummyCost,
                                                                              tempServedRequests2, tMatrix,
                                                                              allRequests, numOfRequests)

                                # if a feasible solution was found
                                if x2PrimeCheckRoute2 != x2PrimeTemp[DN]:
                                    # update all routes
                                    [x2PrimeTemp, x2PrimeRequestTemp, x2PrimeLoadTemp, x2PrimeTimeWastedTemp,
                                     x2PrimeTimeTemp, x2PrimeRewardTemp, x2PrimeCostTemp, x2PrimeFinalRewardTemp,
                                     x2PrimeFinalCostTemp, x2PrimePartiallyServedTemp, x2PrimeServedTemp,
                                     tempPartiallyServedRequestsAllVehicles, tempServedRequestsAll, isItPossible] = \
                                        updateAllRoutes(x2PrimeTemp, x2PrimeRequestTemp, xPrimeTime, tMatrix,
                                                        allRequests)

                                    # check whether profits have increased
                                    x2PrimeProfitTemp = sum(x2PrimeRewardTemp) - sum(x2PrimeCostTemp)

                                    if x2PrimeProfitTemp > bestProfit and isItPossible:  # if we found an improvement
                                        # replace the best
                                        bestProfit = x2PrimeProfitTemp
                                        bestRoute = copy.deepcopy(x2PrimeTemp)
                                        bestRouteRequest = copy.deepcopy(x2PrimeRequestTemp)

    return bestRoute, bestRouteRequest, bestProfit


def findBestSwap(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeServedAll, xPrimeReward, xPrimeCost,
                 xPrimeRewardTotal, xPrimeCostTotal, tMatrix, allRequests, numOfRequests):

    bestProfit = xPrimeRewardTotal - xPrimeCostTotal
    bestRoute = copy.deepcopy(xPrime)
    bestRouteRequest = copy.deepcopy(xPrimeRequest)
    for DN in range(len(xPrime)):
        for i in range(len(xPrimeServed[DN])):
            if xPrimeServed[DN][i] != driverNo[DN]:
                consideredRequest = xPrimeServed[DN][i]
                x2Prime = copy.deepcopy(xPrime)
                x2PrimeRequest = copy.deepcopy(xPrimeRequest)
                x2PrimeCost = copy.deepcopy(xPrimeCost)
                x2PrimeReward = copy.deepcopy(xPrimeReward)

                # delete request from route
                [x2Prime[DN], x2PrimeRequest[DN]] = deleteRequestFromRoute(x2Prime[DN], x2PrimeRequest[DN],
                                                                           consideredRequest)

                # try to insert any other request in route
                [x2Prime[DN], x2PrimeRequest[DN], x2PrimeCost[DN], x2PrimeReward[DN]] = \
                    findBestInsertionSV(x2Prime[DN], x2PrimeRequest[DN], xPrimeTime[DN], xPrimeReward[DN],
                                        xPrimeCost[DN], xPrimeServedAll, tMatrix, allRequests, numOfRequests)

                x2PrimeProfit = sum(x2PrimeReward) - sum(x2PrimeCost) # update the profit

                if x2PrimeProfit > bestProfit:  # if we found an improvement
                    # replace the best
                    bestProfit = x2PrimeProfit
                    bestRoute = copy.deepcopy(x2Prime)
                    bestRouteRequest = copy.deepcopy(x2PrimeRequest)

    return bestRoute, bestRouteRequest, bestProfit


def findBestInsertionSV(currentRoute, currentRequests, currentRouteTime, totalReward, totalCost, currentServed,
                        tMatrix, allRequests, totalNoOfRequests):
    listOfRequestNo = list(range(1, totalNoOfRequests + 1))
    listOfUnservedRequestNoTemp = list(set(listOfRequestNo) - set(currentServed))
    #  listOfUnservedRequestNo = []
    if percentImprovement:  # consider a subset of requests
        if len(listOfUnservedRequestNoTemp) > totalNoOfRequests * percentNumber:
            listOfUnservedRequestNo = random.sample(listOfUnservedRequestNoTemp, int(totalNoOfRequests / 4))
        else:
            listOfUnservedRequestNo = listOfUnservedRequestNoTemp[:]
    else:
        listOfUnservedRequestNo = listOfUnservedRequestNoTemp[:]

    bestRoute = currentRoute[:]
    bestRouteRequests = currentRequests[:]
    bestReward = totalReward
    bestCost = totalCost

    for k in range(len(listOfUnservedRequestNo)):
        requestToTest = listOfUnservedRequestNo[k]  # currently considered request

        for i in range(1, len(currentRoute)):
            proposedRoute = currentRoute[:]
            proposedRouteRequests = currentRequests[:]
            # insert pickup node at i'th place
            proposedRoute.insert(i, allRequests[requestToTest - 1].pickupNode)
            proposedRouteRequests.insert(i, requestToTest)

            for j in range(i+1, len(proposedRoute)):
                tempProposedRoute = proposedRoute[:]
                tempProposedRouteRequests = proposedRouteRequests[:]
                # insert delivery node at j'th place
                tempProposedRoute.insert(j, allRequests[requestToTest - 1].deliveryNode)
                tempProposedRouteRequests.insert(j, requestToTest)

                [tempProposedRoute, tempProposedRouteRequests, tempRouteLoad, tempRouteTimeWasted, tempRouteTime,
                 tempFinalReward, tempFinalCost, tempPartiallyServedRequests, tempServedRequests, isItPossible] = \
                    updateRouteInfo(tempProposedRoute, tempProposedRouteRequests, currentRouteTime, tMatrix,
                                    allRequests)
                if isItPossible and (tempFinalReward - tempFinalCost > bestReward - bestCost):
                    # replace the best
                    bestRoute = tempProposedRoute[:]
                    bestRouteRequests = tempProposedRouteRequests[:]
                    bestCost = tempFinalCost
                    bestReward = tempFinalReward

    return bestRoute, bestRouteRequests, bestCost, bestReward


def findBestInsertion(currentRoute, currentRequests, currentRouteTime, totalReward, totalCost, currentServedAll,
                      tMatrix, allRequests, totalNoOfRequests):
    listOfRequestNo = list(range(1, totalNoOfRequests + 1))
    listOfUnservedRequestNoTemp = list(set(listOfRequestNo) - set(currentServedAll))

    if percentImprovement:  # consider a subset of requests
        if len(listOfUnservedRequestNoTemp) > totalNoOfRequests * percentNumber:
            listOfUnservedRequestNo = random.sample(listOfUnservedRequestNoTemp, int(totalNoOfRequests / 4))
        else:
            listOfUnservedRequestNo = listOfUnservedRequestNoTemp[:]
    else:
        listOfUnservedRequestNo = listOfUnservedRequestNoTemp[:]

    # best insertion is based on the request that gives the highest reward
    bestRoute = copy.deepcopy(currentRoute)
    bestRouteRequests = copy.deepcopy(currentRequests)
    bestReward = totalReward
    bestCost = totalCost
    for DN in range(len(currentRoute)):  # for each driver
        for k in range(len(listOfUnservedRequestNo)):
            requestToTest = listOfUnservedRequestNo[k]  # currently considered request

            for i in range(1, len(currentRoute[DN])):
                proposedRoute = copy.deepcopy(currentRoute)
                proposedRouteRequests = copy.deepcopy(currentRequests)
                # insert pickup node at i'th place
                proposedRoute[DN].insert(i, allRequests[requestToTest - 1].pickupNode)
                proposedRouteRequests[DN].insert(i, requestToTest)

                for j in range(i+1, len(proposedRoute[DN])):
                    tempProposedRoute = copy.deepcopy(proposedRoute)
                    tempProposedRouteRequests = copy.deepcopy(proposedRouteRequests)
                    # insert delivery node at j'th place
                    tempProposedRoute[DN].insert(j, allRequests[requestToTest - 1].deliveryNode)
                    tempProposedRouteRequests[DN].insert(j, requestToTest)

                    [tempProposedRoute, tempProposedRouteRequests, tempRouteLoad, tempRouteTimeWasted, tempRouteTime,
                     tempRouteReward, tempRouteCost, tempFinalReward, tempFinalCost, tempPartiallyServedRequests,
                     tempServedRequests, tempPartiallyServedRequestsAllVehicles, tempServedRequestsAll,
                     isItPossible] = updateAllRoutes(tempProposedRoute, tempProposedRouteRequests, currentRouteTime,
                                                     tMatrix, allRequests)

                    if isItPossible and (tempFinalReward - tempFinalCost > bestReward - bestCost):
                        # replace best
                        bestRoute = copy.deepcopy(tempProposedRoute)
                        bestRouteRequests = copy.deepcopy(tempProposedRouteRequests)
                        bestCost = tempFinalCost
                        bestReward = tempFinalReward

    return bestRoute, bestRouteRequests, bestCost, bestReward

def findRandomInsertionSV(currentRoute, currentRequests, currentRouteTime, totalReward, totalCost, currentServed,
                        tMatrix, allRequests, totalNoOfRequests, dummyRoute):
    listOfRequestNo = list(range(1, totalNoOfRequests + 1))
    listOfUnservedRequestNo = list(set(listOfRequestNo) - set(currentServed))
    bestRoute = currentRoute[:]
    bestRouteRequests = currentRequests[:]
    bestReward = totalReward
    bestCost = totalCost
    bestRewardRatio = 1
    random.shuffle(listOfUnservedRequestNo)
    for k in range(len(listOfUnservedRequestNo)):
        requestToTest = listOfUnservedRequestNo[k]

        iter1 = list(range(1, len(currentRoute)))
        random.shuffle(iter1)
        for i in iter1:
            proposedRoute = currentRoute[:]
            proposedRouteRequests = currentRequests[:]
            proposedRoute.insert(i, allRequests[requestToTest - 1].pickupNode)
            proposedRouteRequests.insert(i, requestToTest)

            iter2 = list(range(i+1, len(proposedRoute)))
            random.shuffle(iter2)
            for j in iter2:
                tempProposedRoute = proposedRoute[:]
                tempProposedRoute.insert(j, allRequests[requestToTest - 1].deliveryNode)
                tempProposedRouteRequests = proposedRouteRequests[:]
                tempProposedRouteRequests.insert(j, requestToTest)

                [tempProposedRoute, tempProposedRouteRequests, tempRouteLoad, tempRouteTimeWasted, tempRouteTime,
                 tempFinalReward, tempFinalCost, tempPartiallyServedRequests, tempServedRequests, isItPossible] = \
                    updateRouteInfo(tempProposedRoute, tempProposedRouteRequests, currentRouteTime, tMatrix,
                                    allRequests)
                if isItPossible and tempProposedRoute != dummyRoute:
                    return tempProposedRoute, tempProposedRouteRequests, tempRouteTime, tempServedRequests, requestToTest

    return currentRoute, currentRequests, currentRouteTime, currentServed, -1

def findRandomMove(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, currentDriverNo, xPrimeReward, xPrimeCost,
                 xPrimeRewardTotal, xPrimeCostTotal, tMatrix, allRequests, numOfRequests):
    bestProfit = xPrimeRewardTotal - xPrimeCostTotal
    bestRoute = copy.deepcopy(xPrime)
    bestRouteRequest = copy.deepcopy(xPrimeRequest)

    DN = currentDriverNo
    random.shuffle(xPrimeServed[DN])
    for i in range(len(xPrimeServed[DN])):  # for each request served in that route
        if xPrimeServed[DN][i] != driverNo[DN]:  # driver's request must not be moved
            # find where the request's nodes are located
            consideredRequest = xPrimeServed[DN][i]
            x2Prime = copy.deepcopy(xPrime)
            x2PrimeRequest = copy.deepcopy(xPrimeRequest)
            x2PrimeCost = copy.deepcopy(xPrimeCost)
            x2PrimeReward = copy.deepcopy(xPrimeReward)

            # remove the request from the route
            [x2Prime[DN], x2PrimeRequest[DN]] = deleteRequestFromRoute(x2Prime[DN], x2PrimeRequest[DN],
                                                                       consideredRequest)
            # create a temporary variable that indicates that the removed request is the only unserved request
            tempServedRequests = list(range(1, numOfRequests + 1))
            del tempServedRequests[xPrimeServed[DN][i] - 1]

            theDrivers = list(range(0, len(xPrime)))
            random.shuffle(theDrivers)
            for j in theDrivers:  # for each route
                if DN != j:  # do not consider the current route
                    x2PrimeTemp = copy.deepcopy(x2Prime)
                    x2PrimeRequestTemp = copy.deepcopy(x2PrimeRequest)
                    x2PrimeCostTemp = x2PrimeCost[:]
                    x2PrimeRewardTemp = xPrimeReward[:]
                    x2PrimeCheckRoute = x2Prime[j][:]

                    dummyReward = 0
                    dummyCost = float("inf")

                    # try to add the removed request inside another route
                    [x2PrimeTemp[j], x2PrimeRequestTemp[j], x2PrimeTimeTemp, dummyServed, requestToTest] =\
                        findRandomInsertionSV(x2PrimeTemp[j], x2PrimeRequestTemp[j], xPrimeTime[j], dummyReward,
                                            dummyCost, tempServedRequests, tMatrix, allRequests, numOfRequests, x2PrimeTemp[j])

                    # if it managed to add the request without violating any constraints
                    if x2PrimeCheckRoute != x2PrimeTemp[j]:
                        # update the information on all routes
                        [x2PrimeTemp, x2PrimeRequestTemp, x2PrimeLoadTemp, x2PrimeTimeWastedTemp,
                         x2PrimeTimeTemp, x2PrimeRewardTemp, x2PrimeCostTemp, x2PrimeFinalRewardTemp,
                         x2PrimeFinalCostTemp, x2PrimePartiallyServedTemp, x2PrimeServedTemp,
                         tempPartiallyServedRequestsAllVehicles, tempServedRequestsAll, isItPossible] = \
                            updateAllRoutes(x2PrimeTemp, x2PrimeRequestTemp, xPrimeTime, tMatrix, allRequests)

                        if isItPossible:
                            return x2PrimeTemp, x2PrimeRequestTemp, x2PrimeTimeTemp, x2PrimeServedTemp, True

    return xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, False


def findRandomRouteSwap(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeReward, xPrimeCost,
                      tMatrix, allRequests, numOfRequests, currentDriverNo):

    bestRoute = copy.deepcopy(xPrime)
    bestRouteRequest = copy.deepcopy(xPrimeRequest)

    DN = currentDriverNo
    random.shuffle(xPrimeServed[DN])
    for i in range(len(xPrimeServed[DN])):  # for each request served in that route
        if xPrimeServed[DN][i] != driverNo[DN]:  # driver's request must not be replaced
            # find where the considered request is located in current route
            consideredRequest = xPrimeServed[DN][i]
            x2Prime = copy.deepcopy(xPrime)
            x2PrimeRequest = copy.deepcopy(xPrimeRequest)
            x2PrimeCost = copy.deepcopy(xPrimeCost)
            x2PrimeReward = copy.deepcopy(xPrimeReward)
            # remove request from the route
            [x2Prime[DN], x2PrimeRequest[DN]] = deleteRequestFromRoute(x2Prime[DN], x2PrimeRequest[DN],
                                                                       consideredRequest)
            # create a temporary variable that indicates that the removed request is the only unserved one
            tempServedRequests = list(range(1, numOfRequests + 1))
            del tempServedRequests[xPrimeServed[DN][i] - 1]

            otherDrivers = list(range(0, len(xPrime)))
            del otherDrivers[DN]
            random.shuffle(otherDrivers)
            for j in otherDrivers:  # for the rest of the drivers/routes
                for m in range(len(xPrimeServed[j])):  # for each request served in the next route
                    if xPrimeServed[j][m] != driverNo[j]:  # driver's request must never be replaced
                        # find where the considered request is located
                        consideredRequest2 = xPrimeServed[j][m]
                        x2PrimeTemp = copy.deepcopy(x2Prime)
                        x2PrimeRequestTemp = copy.deepcopy(x2PrimeRequest)
                        x2PrimeCostTemp = x2PrimeCost[:]
                        x2PrimeRewardTemp = xPrimeReward[:]
                        # remove the considered request
                        [x2PrimeTemp[j], x2PrimeRequestTemp[j]] = \
                            deleteRequestFromRoute(x2PrimeTemp[j], x2PrimeRequestTemp[j], consideredRequest2)

                        x2PrimeCheckRoute = x2PrimeTemp[j][:]
                        dummyReward = 0
                        dummyCost = float("inf")

                        # try to insert the removed request from the previous route into this one
                        [x2PrimeTemp[j], x2PrimeRequestTemp[j], x2PrimeTimeTemp, dummyServed, requestToTest] =\
                            findRandomInsertionSV(x2PrimeTemp[j], x2PrimeRequestTemp[j], xPrimeTime[j], dummyReward,
                                                  dummyCost, tempServedRequests, tMatrix, allRequests, numOfRequests,
                                                  x2PrimeTemp[j])

                        # if a feasible insertion was found
                        if x2PrimeCheckRoute != x2PrimeTemp[j]:
                            x2PrimeCheckRoute2 = x2PrimeTemp[DN][:]
                            tempServedRequests2 = list(range(1, numOfRequests + 1))
                            del tempServedRequests2[xPrimeServed[j][m] - 1]

                            # try to insert the removed request from this route into the previous one
                            [x2PrimeTemp[DN], x2PrimeRequestTemp[DN], x2PrimeTimeTemp, dummyServed, requestToTest] =\
                                findRandomInsertionSV(x2PrimeTemp[DN], x2PrimeRequestTemp[DN], xPrimeTime[DN],
                                                      dummyReward, dummyCost, tempServedRequests2, tMatrix, allRequests,
                                                      numOfRequests, x2PrimeTemp[DN])

                            # if a feasible solution was found
                            if x2PrimeCheckRoute2 != x2PrimeTemp[DN]:
                                # update all routes
                                [x2PrimeTemp, x2PrimeRequestTemp, x2PrimeLoadTemp, x2PrimeTimeWastedTemp,
                                 x2PrimeTimeTemp, x2PrimeRewardTemp, x2PrimeCostTemp, x2PrimeFinalRewardTemp,
                                 x2PrimeFinalCostTemp, x2PrimePartiallyServedTemp, x2PrimeServedTemp,
                                 tempPartiallyServedRequestsAllVehicles, tempServedRequestsAll, isItPossible] = \
                                    updateAllRoutes(x2PrimeTemp, x2PrimeRequestTemp, xPrimeTime, tMatrix,
                                                    allRequests)
                                if isItPossible:
                                    return x2PrimeTemp, x2PrimeRequestTemp, x2PrimeTimeTemp, x2PrimeServedTemp, True

    return xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, False


def findRandomSwap(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeServedAll, xPrimeReward, xPrimeCost,
                 tMatrix, allRequests, numOfRequests, currentDriverNo):

    bestRoute = copy.deepcopy(xPrime)
    bestRouteRequest = copy.deepcopy(xPrimeRequest)
    DN = currentDriverNo
    random.shuffle(xPrimeServed[DN])
    for i in range(len(xPrimeServed[DN])):
        if xPrimeServed[DN][i] != driverNo[DN]:
            requestToRemove = xPrimeServed[DN][i]
            consideredRequest = xPrimeServed[DN][i]
            x2Prime = copy.deepcopy(xPrime)
            x2PrimeRequest = copy.deepcopy(xPrimeRequest)
            x2PrimeCost = copy.deepcopy(xPrimeCost)
            x2PrimeReward = copy.deepcopy(xPrimeReward)
            x2PrimeTime = copy.deepcopy(xPrimeTime)
            x2PrimeServed = copy.deepcopy(xPrimeServed)

            # delete request from route
            [x2Prime[DN], x2PrimeRequest[DN]] = deleteRequestFromRoute(x2Prime[DN], x2PrimeRequest[DN],
                                                                       consideredRequest)

            # find a random feasible insertion
            [x2Prime[DN], x2PrimeRequest[DN], x2PrimeCost[DN], x2PrimeReward[DN], requestToTest] = \
                findRandomInsertionSV(x2Prime[DN], x2PrimeRequest[DN], xPrimeTime[DN], xPrimeReward[DN],
                                    xPrimeCost[DN], xPrimeServedAll, tMatrix, allRequests, numOfRequests, x2Prime[DN])

            [x2Prime[DN], x2PrimeRequest[DN], tempRouteLoad, tempRouteTimeWasted, x2PrimeTime[DN],
             tempFinalReward, tempFinalCost, tempPartiallyServedRequests, x2PrimeServed[DN], isItPossible] = \
                updateRouteInfo(x2Prime[DN], x2PrimeRequest[DN], x2PrimeTime[DN], tMatrix, allRequests)
            if isItPossible and requestToTest != -1:  # if we found an improvement
                xPrimeServedAll.remove(requestToRemove)
                xPrimeServedAll.append(requestToTest)
                return x2Prime, x2PrimeRequest, x2PrimeTime, x2PrimeServed, xPrimeServedAll, True

    return xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeServedAll, False


def findRandomIntraSwap(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeServedAll, xPrimeReward, xPrimeCost,
                      tMatrix, allRequests, numOfRequests, currentDriverNo):
    bestRoute = copy.deepcopy(xPrime)
    bestRouteRequest = copy.deepcopy(xPrimeRequest)

    DN = currentDriverNo

    random.shuffle(xPrimeServed[DN])
    for i in range(len(xPrimeServed[DN])):  # for each request served
        if xPrimeServed[DN][i] != driverNo[DN]:  # driver start and end nodes must never be changed
            # find where the nodes of the considered request are located in the route
            consideredRequest= xPrimeServed[DN][i]
            x2Prime = copy.deepcopy(xPrime)
            x2PrimeRequest = copy.deepcopy(xPrimeRequest)
            x2PrimeCost = copy.deepcopy(xPrimeCost)
            x2PrimeReward = copy.deepcopy(xPrimeReward)
            x2PrimeTime = copy.deepcopy(xPrimeTime)
            x2PrimeServed = copy.deepcopy(xPrimeServed)
            tempRoute = x2Prime[DN][:]

            # remove the request from the route
            [x2Prime[DN], x2PrimeRequest[DN]] = deleteRequestFromRoute(x2Prime[DN], x2PrimeRequest[DN],
                                                                       consideredRequest)

            # create a temporary variable that allows only the removed request to be unserved
            tempServedRequests = list(range(1, numOfRequests + 1))
            del tempServedRequests[xPrimeServed[DN][i] - 1]

            totalReward = 0
            totalCost = float("inf")

            # find best insertion will only try to add the removed request
            [x2Prime[DN], x2PrimeRequest[DN], x2PrimeCost[DN], x2PrimeReward[DN], requestToTest] = \
                findRandomInsertionSV(x2Prime[DN], x2PrimeRequest[DN], xPrimeTime[DN], xPrimeReward[DN],
                                      xPrimeCost[DN], tempServedRequests, tMatrix, allRequests, numOfRequests,
                                      tempRoute)

            [x2Prime[DN], x2PrimeRequest[DN], tempRouteLoad, tempRouteTimeWasted, x2PrimeTime[DN],
             tempFinalReward, tempFinalCost, tempPartiallyServedRequests, x2PrimeServed[DN], isItPossible] = \
                updateRouteInfo(x2Prime[DN], x2PrimeRequest[DN], x2PrimeTime[DN], tMatrix, allRequests)

            if isItPossible and requestToTest != -1 and tempRoute != x2Prime[DN]:
                return x2Prime, x2PrimeRequest, x2PrimeTime, True

    return xPrime, xPrimeRequest, xPrimeTime, False


def VNS(currentRoute, currentRouteRequest, currentRouteTime, currentRouteLoad, currentRouteServed, currentReward,
        currentCost, tMatrix, allRequests, numOfRequests, dataOfImages):

    k = 2
    noOfNonImproving = 0
    bestRoute = copy.deepcopy(currentRoute)
    bestRouteRequest = copy.deepcopy(currentRouteRequest)
    bestRequestServed = copy.deepcopy(currentRouteServed)
    bestReward = currentReward[:]
    bestCost = currentCost[:]
    xTime = copy.deepcopy(currentRouteTime)
    noOfGS = 0
    counterGSserved = []
    counterGS = []
    counterGSprofit = []
    startTimeVNS = time.clock()
    TTFBS = 0

    while noOfNonImproving < 300:
        k = min(2, kmax)
        while k <= kmax:
            x = copy.deepcopy(bestRoute)
            xRequest = copy.deepcopy(bestRouteRequest)
            xServedRequests = copy.deepcopy(bestRequestServed)
            xReward = bestReward
            xCost = bestCost
            xServedRequestsAll = []
            for DN in range(len(xServedRequests)):
                for j in range(len(xServedRequests[DN])):
                    xServedRequestsAll.append(xServedRequests[DN][j])
            noOfMoves = 0
            tabuList = []
            dummyRequests = copy.deepcopy(xRequest)
            tabuList.append(dummyRequests)

            start_time_GS = time.clock()
            while noOfMoves < k:
                oper = random.randint(1, 6)
                routeNo = random.randint(0, len(driverNo) - 1)

                if (oper == 1):  # make random feasible insert
                    [xTemp2, xRequestTemp2, xTimeTemp, xServedRequestsTemp2, addedRequest] = \
                        findRandomInsertionSV(x[routeNo], xRequest[routeNo], xTime[routeNo], 0, 0, xServedRequestsAll,
                                              tMatrix, allRequests, numOfRequests, x[routeNo])

                    if addedRequest != -1 and not any(xk == xRequestTemp2 for xk in tabuList):

                        noOfMoves += 1
                        x[routeNo] = xTemp2[:]
                        xRequest[routeNo] = xRequestTemp2[:]
                        tabuList.append(copy.deepcopy(xRequest))

                elif (oper == 2):  # make random feasible move
                    if len(xRequest[routeNo]) > 2:
                        [xTemp, xRequestTemp, xTimeTemp, xServedRequestsTemp, isItChanged] =\
                            findRandomMove(x, xRequest, xTime, xServedRequests, routeNo, xReward, xCost, sum(xReward),
                                           sum(xCost), tMatrix, allRequests, numOfRequests)

                        if isItChanged and not any(xk == xRequestTemp for xk in tabuList):

                            noOfMoves += 1
                            x = copy.deepcopy(xTemp)
                            xRequest = copy.deepcopy(xRequestTemp)
                            tabuList.append(copy.deepcopy(xRequest))

                elif (oper == 3):  # make random feasible inter-swap
                    if len(xRequest[routeNo]) > 2:
                        [xTemp, xRequestTemp, xTimeTemp, xServedRequestsTemp, isItChanged] =\
                            findRandomRouteSwap(x, xRequest, xTime, xServedRequests, xReward, xCost, tMatrix,
                                                allRequests, numOfRequests, routeNo)

                        if isItChanged and not any(xk == xRequestTemp for xk in tabuList):

                            noOfMoves += 1
                            x = copy.deepcopy(xTemp)
                            xRequest = copy.deepcopy(xRequestTemp)
                            tabuList.append(copy.deepcopy(xRequest))

                elif (oper == 4):  # make random feasible swap
                    if len(xRequest[routeNo]) > 2:
                        [xTemp, xRequestTemp, xTimeTemp, xServedRequestsTemp, xServedRequestsAllTemp, isItChanged] =\
                            findRandomSwap(x, xRequest, xTime, xServedRequests, xServedRequestsAll, xReward, xCost,
                                           tMatrix, allRequests, numOfRequests, routeNo)

                        if isItChanged and not any(xk == xRequestTemp for xk in tabuList):

                            noOfMoves += 1
                            x = copy.deepcopy(xTemp)
                            xRequest = copy.deepcopy(xRequestTemp)
                            tabuList.append(copy.deepcopy(xRequest))

                elif (oper == 5):  # make feasible intra-swap
                    if len(xRequest[routeNo]) > 2:
                        [xTemp, xRequestTemp, xTimeTemp, isItChanged] =\
                            findRandomIntraSwap(x, xRequest, xTime, xServedRequests, xServedRequestsAll, xReward, xCost,
                                                tMatrix, allRequests, numOfRequests, routeNo)

                        if isItChanged and not any(xk == xRequestTemp for xk in tabuList):

                            noOfMoves += 1
                            x = copy.deepcopy(xTemp)
                            xRequest = copy.deepcopy(xRequestTemp)
                            tabuList.append(copy.deepcopy(xRequest))

                elif (oper == 6):  # make deletion
                    if len(xRequest[routeNo]) > 2:
                        xTemp = copy.deepcopy(x)
                        xRequestTemp = copy.deepcopy(xRequest)
                        requestsToChooseFrom = xServedRequests[routeNo][:]
                        requestsToChooseFrom.remove(driverNo[routeNo])
                        requestsToRemove = random.sample(requestsToChooseFrom, 1)
                        consideredRequest = requestsToRemove[0]
                        # delete the nodes serving that request
                        [xTemp[routeNo], xRequestTemp[routeNo]] =\
                            deleteRequestFromRoute(xTemp[routeNo], xRequestTemp[routeNo], consideredRequest)

                        [x, xRequest, xLoad, xTimeWasted, xTime, xReward, xCost, xFinalReward, xFinalCost,
                         xPartiallyServedRequests, xServedRequests, xPartiallyServedRequestsAllVehicles,
                         xServedRequestsAll, isItPossibleTemp] = \
                            updateAllRoutes(x, xRequest, currentRouteTime, tMatrix, allRequests)

                        if isItPossibleTemp and not any(xk == xRequestTemp for xk in tabuList):

                            noOfMoves += 1
                            x = copy.deepcopy(xTemp)
                            xRequest = copy.deepcopy(xRequestTemp)
                            tabuList.append(copy.deepcopy(xRequest))
                [x, xRequest, xLoad, xTimeWasted, xTime, xReward, xCost, xFinalReward, xFinalCost,
                 xPartiallyServedRequests,
                 xServedRequests, xPartiallyServedRequestsAllVehicles, xServedRequestsAll, isItPossible] = \
                    updateAllRoutes(x, xRequest, currentRouteTime, tMatrix, allRequests)

                if time.clock() - start_time_GS > 5:  # empty the tabu list if GS takes longer than 5 seconds
                    tabuList = []

            [x, xRequest, xLoad, xTime, xServedRequests, xServedRequestsAll, xReward, xCost, dataOfImages] = \
                localSearch(x, xRequest, xTime, xLoad, xReward, xCost, xServedRequests, xServedRequestsAll, tMatrix,
                            allRequests, numOfRequests, dataOfImages)

            xProfit = sum(xReward) - sum(xCost)
            bestProfit = sum(bestReward) - sum(bestCost)

            noOfGS += 1

            matches = [k for k, x in enumerate(counterGSserved) if x == xServedRequestsAll]
            if matches == []:
                counterGSserved.append(xServedRequestsAll)
                counterGSprofit.append(round(xProfit, 2))
                counterGS.append(1)
            else:
                counterGS[matches[0]] += 1

            if xProfit > bestProfit:
                # a better solution is found
                k = 2
                noOfNonImproving = 0
                bestRoute = copy.deepcopy(x)
                bestRouteRequest = copy.deepcopy(xRequest)
                bestRequestServed = copy.deepcopy(xServedRequests)
                bestReward = xReward[:]
                bestCost = xCost[:]
                bestProfit = sum(bestReward) - sum(bestCost)
                TTFBS = time.clock() - startTimeVNS
            else:
                k += 1

            if time.clock() - startTimeVNS > maxCPUtime:
                break

        noOfNonImproving += 1
        if time.clock() - startTimeVNS > maxCPUtime:
            break

    return bestRoute, bestRouteRequest, noOfGS, TTFBS


def localSearch(xPrime, xPrimeRequest, xPrimeTime, xPrimeLoad, xPrimeReward, xPrimeCost, xPrimeServed, xPrimeServedAll,
                tMatrix, allRequests, numOfRequests, dataOfImages):

    operators = ["insert", "move", "inter-swap", "swap", "intra-swap", "deletion"]
    level = 1
    level_max = len(operators)
    noOfImprovements = 0
    xPrimeTimeWasted = []
    x2FinalCost = 0
    x2FinalReward = 0

    while level <= level_max:
        # reset x2Prime values
        x2Prime = []
        x2PrimeRequest = []
        x2PrimeTime = []
        x2PrimeLoad = []
        x2PrimeServedRequests = []
        x2PrimeCost = 0
        x2PrimeReward = 0
        isItPossible = False

        xPrimeFinalReward = sum(xPrimeReward)
        xPrimeFinalCost = sum(xPrimeCost)

        if level == 1:  # insert operator
            [x2Prime, x2PrimeRequest, x2FinalCost, x2FinalReward] =\
                            findBestInsertion(xPrime, xPrimeRequest, xPrimeTime, xPrimeFinalReward, xPrimeFinalCost,
                                              xPrimeServedAll, tMatrix, allRequests, numOfRequests)

            [x2Prime, x2PrimeRequest, x2PrimeLoad, x2PrimeTimeWasted, x2PrimeTime, x2PrimeReward, x2PrimeCost,
             x2FinalReward, x2FinalCost, x2PrimePartiallyServed, x2PrimeServedRequests,
             x2PartiallyServedRequestsAllVehicles, x2ServedRequestsAll, isItPossible] = \
                updateAllRoutes(x2Prime, x2PrimeRequest, xPrimeTime, tMatrix, allRequests)

        elif level == 2:  # move operator

            [x2Prime, x2PrimeRequest, x2PrimeProfit] =\
                            findBestMove(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeReward, xPrimeCost,
                                         xPrimeFinalReward, xPrimeFinalCost, tMatrix, allRequests, numOfRequests)

            [x2Prime, x2PrimeRequest, x2PrimeLoad, x2PrimeTimeWasted, x2PrimeTime, x2PrimeReward, x2PrimeCost,
             x2FinalReward, x2FinalCost, x2PrimePartiallyServed, x2PrimeServedRequests,
             x2PartiallyServedRequestsAllVehicles, x2ServedRequestsAll, isItPossible] = \
                                            updateAllRoutes(x2Prime, x2PrimeRequest, xPrimeTime, tMatrix, allRequests)

        elif level == 3:  # inter-swap operator

            [x2Prime, x2PrimeRequest, x2PrimeProfit] =\
                            findBestRouteSwap(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeReward, xPrimeCost,
                                              xPrimeFinalReward, xPrimeFinalCost, tMatrix, allRequests, numOfRequests)

            [x2Prime, x2PrimeRequest, x2PrimeLoad, x2PrimeTimeWasted, x2PrimeTime, x2PrimeReward, x2PrimeCost,
             x2FinalReward, x2FinalCost, x2PrimePartiallyServed, x2PrimeServedRequests,
             x2PartiallyServedRequestsAllVehicles, x2ServedRequestsAll, isItPossible] = \
                                            updateAllRoutes(x2Prime, x2PrimeRequest, xPrimeTime, tMatrix, allRequests)

        elif level == 4:  # swap operator

            [x2Prime, x2PrimeRequest, x2PrimeProfit] =\
                            findBestSwap(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeServedAll, xPrimeReward,
                                         xPrimeCost, xPrimeFinalReward, xPrimeFinalCost, tMatrix, allRequests,
                                         numOfRequests)

            [x2Prime, x2PrimeRequest, x2PrimeLoad, x2PrimeTimeWasted, x2PrimeTime, x2PrimeReward, x2PrimeCost,
             x2FinalReward, x2FinalCost, x2PrimePartiallyServed, x2PrimeServedRequests,
             x2PartiallyServedRequestsAllVehicles, x2ServedRequestsAll, isItPossible] = \
                                            updateAllRoutes(x2Prime, x2PrimeRequest, xPrimeTime, tMatrix, allRequests)

        elif level == 5:  # intra-swap operator

            [x2Prime, x2PrimeRequest] = findBestIntraSwap(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed, xPrimeReward,
                                                          xPrimeCost, tMatrix, allRequests, numOfRequests)

            [x2Prime, x2PrimeRequest, x2PrimeLoad, x2PrimeTimeWasted, x2PrimeTime, x2PrimeReward, x2PrimeCost,
             x2FinalReward, x2FinalCost, x2PrimePartiallyServed, x2PrimeServedRequests,
             x2PartiallyServedRequestsAllVehicles, x2ServedRequestsAll, isItPossible] = \
                                            updateAllRoutes(x2Prime, x2PrimeRequest, xPrimeTime, tMatrix, allRequests)

        elif level == 6:  # delete operator

            [x2Prime, x2PrimeRequest, x2PrimeProfit] = findBestDeletion(xPrime, xPrimeRequest, xPrimeTime, xPrimeServed,
                                                                  xPrimeReward, xPrimeCost, tMatrix, allRequests)

            [x2Prime, x2PrimeRequest, x2PrimeLoad, x2PrimeTimeWasted, x2PrimeTime, x2PrimeReward, x2PrimeCost,
             x2FinalReward, x2FinalCost, x2PrimePartiallyServed, x2PrimeServedRequests,
             x2PartiallyServedRequestsAllVehicles, x2ServedRequestsAll, isItPossible] = \
                                            updateAllRoutes(x2Prime, x2PrimeRequest, xPrimeTime, tMatrix, allRequests)

        x2PrimeProfit = x2FinalReward - x2FinalCost
        xPrimeProfit = sum(xPrimeReward) - sum(xPrimeCost)
        if x2PrimeProfit > xPrimeProfit and isItPossible:  # if we found an improvement
            xPrime = copy.deepcopy(x2Prime)
            xPrimeRequest = copy.deepcopy(x2PrimeRequest)
            xPrimeTime = copy.deepcopy(x2PrimeTime)
            xPrimeTimeWasted = copy.deepcopy(x2PrimeTimeWasted)
            xPrimeReward = x2PrimeReward[:]
            xPrimeCost = x2PrimeCost[:]
            xPrimeLoad = copy.deepcopy(x2PrimeLoad)
            xPrimeServed = copy.deepcopy(x2PrimeServedRequests)
            xPrimeServedAll = x2ServedRequestsAll[:]
            xPrimeFinalReward = x2FinalReward
            xPrimeFinalCost = x2FinalCost
            level = 1
            xProfitPlot.append([time.clock() - start_time, x2PrimeProfit])

        else:  # if we did not find an improvement, move to the next operator
            level += 1
            xProfitPlot.append([time.clock() - start_time, xPrimeProfit])

    [xPrime, xPrimeRequest, xPrimeLoad, xPrimeTimeWasted, xPrimeTime, xPrimeReward, xPrimeCost, xFinalReward,
     xFinalCost, xPrimePartiallyServed, xPrimeServedRequests, xPartiallyServedRequestsAllVehicles, xServedRequestsAll,
     isItPossible] = updateAllRoutes(xPrime, xPrimeRequest, xPrimeTime, tMatrix, allRequests)
    return xPrime, xPrimeRequest, xPrimeLoad, xPrimeTime, xPrimeServed, xPrimeServedAll, xPrimeReward, xPrimeCost,\
           dataOfImages


def checkPrecedenceConstraintsRoute(currentRoute, currentRequests, allRequests):
    for i in range(len(currentRequests)):
        matches = [k for k, x in enumerate(currentRequests) if x == currentRequests[i]]
        # if delivery node is found prior to the pickup node in the request, precedence constraints are violated
        if currentRoute[matches[0]] == allRequests[currentRequests[i] - 1].deliveryNode:
            return False
    return True


def checkCapacityConstraints(consideredRequest, allowedCapacity, totalLoad, allRequests):
    # we want to make sure that adding the load of the considered node to the vehicle will not exceed capacity
    # load total = load before arriving at node
    # allRequests[ ].load = load at considered Node
    return totalLoad + allRequests[consideredRequest - 1].load <= allowedCapacity


def checkTimeWindowsConstraints(consideredRequest, consideredNode, currentRoute, tMatrix, allRequests, currentTime):
    checked = True  # if this is false, that means that the node considered cannot be added due to time windows
    currTime = currentTime  # this is used as a dummy variable for the total Time of the route
    timeWasted = 0  # this is used as an indication of the total waiting time for a customer in a pickup location
    e1 = 0
    l1 = 0

    travelTime = tMatrix[currentRoute[-1] - 1][consideredNode - 1]  # travel time from last node visited to next node
    currTime += travelTime  # add it to total time

    # find the time windows of the next node to be added to the route
    if consideredNode == allRequests[consideredRequest - 1].pickupNode:  # if considered node is a pickup node
        e1 = allRequests[consideredRequest - 1].pickupTimeWindowE
        l1 = allRequests[consideredRequest - 1].pickupTimeWindowL
    elif consideredNode == allRequests[consideredRequest - 1].deliveryNode:  # if considered node is a delivery node
        e1 = allRequests[consideredRequest - 1].deliveryTimeWindowE
        l1 = allRequests[consideredRequest - 1].deliveryTimeWindowL
    if currTime < e1:  # driver must wait and time is wasted
        timeWasted += e1 - currTime
        checked = True
    elif currTime <= l1:  # on time
        checked = True
    else:  # if time window not reached
        checked = False

    return checked, timeWasted, travelTime

def checkIfRouteIsPossible(currentRoute, consideredRequest, consideredNode, timeTotal, travelTime, timeWasted,
                           timeLimit, tMatrix, driverFinalDestination, allRequests, partialServedRequests,
                           driverNumber):
    # after time windows, capacity and precedence constraints are satisfied for the considered node, we try to add it
    # ensuring that other constraints concerning other nodes (and this one) are not violated

    nodesLeftToServe = []
    dummyPartiallyServedRequests = partialServedRequests[:]
    dummyPartiallyServedRequests.remove(driverNumber)  # request of driver is not to be considered (it's the end point)

    if type(dummyPartiallyServedRequests) != 'NoneType':
        for i in range(len(dummyPartiallyServedRequests)):
            # we check which delivery nodes must be visited to complete the requests
            nodesLeftToServe.append(allRequests[dummyPartiallyServedRequests[i] - 1].deliveryNode)
    if consideredNode != allRequests[consideredRequest - 1].deliveryNode:  # if we are considering a pickup node
        nodesLeftToServe.append(allRequests[consideredRequest - 1].deliveryNode)  # add its corresponding delivery node
        dummyPartiallyServedRequests.append(consideredRequest)

    # all the possible combinations of the unserved delivery nodes
    permuts = list(itertools.permutations(nodesLeftToServe))
    permutsRequests = list(itertools.permutations(dummyPartiallyServedRequests))

    if permuts == [()]:
        return True  # if permuts is empty, then we have no unserved delivery nodes
    else:
        bestTime = float("inf")  # no best Time so far, therefore, just put it to infinity for now

        for i in range(len(permuts)):  # for all permutations
            routeCopy = currentRoute[:]
            routeCopy.append(consideredNode)  # append considered node to the route
            # the new time is the previous total Time summed with the time required to add the considered node
            theTotalTime = timeTotal + travelTime + timeWasted
            dummyTime = theTotalTime

            # placing the current permutation in the sequence to be tested for satisfaction of constraints
            sequence = permuts[i]
            requestSequence = permutsRequests[i]

            # if all of these are true, then route is possible w.r.t. time windows
            timeWindowConstraints = [False] * len(sequence)

            for j in range(len(sequence)):  # for each node in sequence
                lastNodeVisited = routeCopy[-1]  # last node in the route
                # travel time from last node visited to the next node in sequence
                tij = tMatrix[lastNodeVisited - 1][sequence[j] - 1]
                sequenceRequest = requestSequence[j]
                # check if (just) the next node is possible w.r.t. time windows to add to the sequence
                # also calculate the total time wasted in waiting for this delivery (there should not be any)
                [checkTime, theTimeWasted, dummyTravelTime] =\
                                        checkTimeWindowsConstraints(sequenceRequest, sequence[j], routeCopy, tMatrix,
                                                                    allRequests, dummyTime)

                timeWindowConstraints[j] = checkTime

                if checkTime == False:  # if a node is not possible due to time windows
                    break  # break out of the for (j) loop

                dummyTime += tij + theTimeWasted  # adding to the total Time
                routeCopy.append(sequence[j])  # and appending the next node since it is possible w.r.t time windows

            if all(timeWindowConstraints) == True:  # if all nodes in the sequence are possible w.r.t. time windows
                # we must append the driver's destination point to the route
                dummyTime += tMatrix[routeCopy[-1] - 1][driverFinalDestination - 1]
                routeCopy.append(driverFinalDestination)

                if dummyTime < bestTime:  # if we found a better option
                    bestTime = dummyTime  # replace best

        # we only return true if there is a combination (sequence) in which the time limits are not violated
        return bestTime <= timeLimit


def updateRouteInfo(currentRoute, currentRequest, currentRouteTime, tMatrix, allRequests):
    # USED TO UPDATE THE INFORMATION ON A SINGLE ROUTE
    # ------------------------------------------------ INITIALIZATION ------------------------------------------------
    tempRoute = []
    tempRouteRequest = []
    tempRouteLoad = []
    tempRouteTimeWasted = []
    tempRouteTime = []
    tempServedRequests = []
    tempPartiallyServedRequests = []
    tempFinalReward = 0
    tempFinalCost = 0

    tempRoute.append(currentRoute[0])
    tempRouteRequest.append(currentRequest[0])
    tempRouteLoad.append(0)
    tempRouteTimeWasted.append(0)
    tempRouteTime.append(currentRouteTime[0])
    tempPartiallyServedRequests.append(currentRequest[0])
    # ----------------------------------------------------------------------------------------------------------------

    for i in range(1, len(currentRoute)):  # testing all the nodes inside the route
        # CAPACITY
        if allRequests[currentRequest[i] - 1].pickupNode == currentRoute[i]:  # if it is a pickup node
            tempRouteLoad.append(tempRouteLoad[-1] + allRequests[currentRequest[i] - 1].load)  # increase the load
            tempPartiallyServedRequests.append(currentRequest[i])
            if tempRouteLoad[-1] > capacity:  # if capacity constraints have been violated
                # exit the function with isItPossible = False
                return currentRoute, currentRequest, tempRouteLoad, tempRouteTimeWasted, tempRouteTime,\
                       tempFinalReward, tempFinalCost, tempPartiallyServedRequests, tempServedRequests, False

        # capacity constraints cannot be violated if we are at a delivery node
        elif allRequests[currentRequest[i] - 1].deliveryNode == currentRoute[i]:
            tempServedRequests.append(currentRequest[i])
            tempPartiallyServedRequests.remove(currentRequest[i])

            # if request considered is the driver's request, then no reward is given
            if currentRequest[i] == currentRequest[0]:
                tempRouteLoad.append(0)
            else:  # update the load and reward
                tempRouteLoad.append(tempRouteLoad[-1] - allRequests[currentRequest[i] - 1].load)
                tempFinalReward += allRequests[currentRequest[i] - 1].reward  # update the reward

        # TIME WINDOWS
        [timeWindowsCheck, temporaryTimeWasted, timeToTravel] =\
                                    checkTimeWindowsConstraints(currentRequest[i], currentRoute[i], tempRoute, tMatrix,
                                                                allRequests, tempRouteTime[-1])
        tempRouteTimeWasted.append(temporaryTimeWasted)  # update time wasted list

        if timeWindowsCheck:  # if time windows constraints are not violated
            tempRouteTime.append(tempRouteTime[-1] + tempRouteTimeWasted[i] + timeToTravel)  # update time list
        else:  # time windows constraints have been violated
            # exit the function with isItPossible = False
            return currentRoute, currentRequest, tempRouteLoad, tempRouteTimeWasted, tempRouteTime, tempFinalReward,\
                   tempFinalCost, tempPartiallyServedRequests, tempServedRequests, False

        # we are not considering wasted time as a cost
        # cost is calculated to be 10c per minute
        tempFinalCost += 0.0016 * timeToTravel
        tempRoute.append(currentRoute[i])
        tempRouteRequest.append(currentRequest[i])

    return tempRoute, tempRouteRequest, tempRouteLoad, tempRouteTimeWasted, tempRouteTime, tempFinalReward,\
           tempFinalCost, tempPartiallyServedRequests, tempServedRequests, True


def updateAllRoutes(currentRoute, currentRequest, currentRouteTime, tMatrix, allRequests):
    # ------------------------------------------------ INITIALIZATION ------------------------------------------------
    tempRoute = []
    tempRouteRequest = []
    tempRouteLoad = []
    tempRouteTimeWasted = []
    tempRouteTime = []
    tempServedRequests = []
    tempPartiallyServedRequests = []
    tempFinalReward = []
    tempFinalCost = []
    tempServedRequestsAll = []
    tempPartiallyServedRequestsAll = []

    # ----------------------------------------------------------------------------------------------------------------
    for j in range(len(driverNo)):

        tempRoute.append([currentRoute[j][0]])
        tempRouteRequest.append([currentRequest[j][0]])
        tempRouteLoad.append([0])
        tempRouteTimeWasted.append([0])
        tempRouteTime.append([currentRouteTime[j][0]])
        tempPartiallyServedRequests.append([currentRequest[j][0]])
        tempServedRequests.append([])
        tempFinalReward.append(0)
        tempFinalCost.append(0)

        for i in range(1, len(currentRoute[j])):  # testing all the nodes
            # CAPACITY
            if allRequests[currentRequest[j][i] - 1].pickupNode == currentRoute[j][i]:  # if pickup node
                # update the load and partially served requests
                tempRouteLoad[j].append(tempRouteLoad[j][-1] + allRequests[currentRequest[j][i] - 1].load)
                tempPartiallyServedRequests[j].append(currentRequest[j][i])
                if tempRouteLoad[j][-1] > capacity:  # if capacity constraints violated
                    # exit the function with isItPossible = False
                    return currentRoute, currentRequest, tempRouteLoad, tempRouteTimeWasted, tempRouteTime,\
                           tempFinalReward, tempFinalCost, 0, 0, tempPartiallyServedRequests, tempServedRequests,\
                           tempPartiallyServedRequestsAll, tempServedRequestsAll, False
            # capacity constraints cannot be violated if we are at a delivery node
            elif allRequests[currentRequest[j][i] - 1].deliveryNode == currentRoute[j][i]:
                tempServedRequests[j].append(currentRequest[j][i])
                tempPartiallyServedRequests[j].remove(currentRequest[j][i])
                if currentRequest[j][i] == currentRequest[j][0]:
                    tempRouteLoad[j].append(0)
                else:
                    tempRouteLoad[j].append(tempRouteLoad[j][-1] - allRequests[currentRequest[j][i] - 1].load)
                    tempFinalReward[j] += allRequests[currentRequest[j][i] - 1].reward  # update the reward

            # TIME WINDOWS
            [timeWindowsCheck, tempTimeWasted, timeToTravel] =\
                            checkTimeWindowsConstraints(currentRequest[j][i], currentRoute[j][i], tempRoute[j], tMatrix,
                                                        allRequests, tempRouteTime[j][-1])

            tempRouteTimeWasted[j].append(tempTimeWasted)  # update time wasted list
            if timeWindowsCheck:  # if time windows constraints are not violated
                # update time list
                tempRouteTime[j].append(tempRouteTime[j][-1] + tempRouteTimeWasted[j][i] + timeToTravel)
            else:  # time windows constraints violated
                # exit the function with isItPossible = False
                return currentRoute, currentRequest, tempRouteLoad, tempRouteTimeWasted, tempRouteTime,\
                       tempFinalReward, tempFinalCost, 0, 0, tempPartiallyServedRequests, tempServedRequests,\
                       tempPartiallyServedRequestsAll, tempServedRequestsAll, False

            # we are not considering wasted time as a cost
            # cost is calculated to be 10c per minute travelled
            tempFinalCost[j] += 0.0016 * timeToTravel
            tempRoute[j].append(currentRoute[j][i])
            tempRouteRequest[j].append(currentRequest[j][i])

        # check if there are any partially served routes left
        for k in range(len(tempPartiallyServedRequests[j])):
            tempPartiallyServedRequestsAll.append(partiallyServedRequests[j][k])
        # update which routes have been served
        for k in range(len(tempServedRequests[j])):
            tempServedRequestsAll.append(tempServedRequests[j][k])

    # update the final reward
    theFinalReward = 0
    theFinalCost = 0
    for j in range(len(tempFinalReward)):
        theFinalReward += tempFinalReward[j]
    for j in range(len(tempFinalCost)):
        theFinalCost += tempFinalCost[j]

    return tempRoute, tempRouteRequest, tempRouteLoad, tempRouteTimeWasted, tempRouteTime, tempFinalReward,\
           tempFinalCost, theFinalReward, theFinalCost, tempPartiallyServedRequests, tempServedRequests,\
           tempPartiallyServedRequestsAll, tempServedRequestsAll, True


def findANode(totalTime, totalLoad, totalRewardEach, totalCostEach, currentRoute, currentRequest, currentLoad,
              currentTimeWasted, currentRouteTime, allowedCapacity, tMatrix, timeLimit, driverEnd, requestsServed,
              requestsServedAll, partialServedRequests, partialServedRequestsAll, arrayOfNodesToConsider, allRequests,
              driverNumber, listOfNodesSortedByCost, totalReward, totalCost):

    # all the nodes to consider (pickup nodes of the unserved requests and delivery nodes of partially served requests
    pickupDeliveryNodes = [i[1] for i in arrayOfNodesToConsider]
    for i in range(len(listOfNodesSortedByCost)):  # testing all the nodes
        # find the requests that have the closest nodes to the last visited node in the route
        # or uniform random distributed
        matches = [k for k, x in enumerate(pickupDeliveryNodes) if x == listOfNodesSortedByCost[i] + 1]
        for k in range(len(matches)):  # if a match is found
            consideredRequest = arrayOfNodesToConsider[matches[k]][0]  # request No
            consideredNode = arrayOfNodesToConsider[matches[k]][1]  # pickup/delivery node number

            # if the node considered in the request chosen is a pickup node, that it has not been served yet
            if allRequests[consideredRequest - 1].pickupNode == consideredNode and \
                consideredRequest not in requestsServed and \
                consideredRequest not in partialServedRequestsAll:

                # if capacity constraints are not violated
                if checkCapacityConstraints(consideredRequest, allowedCapacity, totalLoad, allRequests):
                    # check if time windows are violated
                    [checkedTimeWindows, wastedTime, travelTime] =\
                                checkTimeWindowsConstraints(consideredRequest, consideredNode, currentRoute, tMatrix,
                                                            allRequests, totalTime)
                    if checkedTimeWindows:  # if time windows not violated
                        # try to find a permutation that allows the node to be served, together with the partially
                        # served requests
                        if checkIfRouteIsPossible(currentRoute, consideredRequest, consideredNode, totalTime,
                                                  travelTime, wastedTime, timeLimit, tMatrix, driverEnd, allRequests,
                                                  partialServedRequests, driverNumber):

                            # if there is a permutation that allows this, we can append the considered node / request
                            partialServedRequests.append(consideredRequest)  # add the request to partially served
                            # add the request to partially served by all vehicles
                            partialServedRequestsAll.append(consideredRequest)
                            # remove the pickup node from nodes to consider, and add the corresponding delivery node
                            # to it
                            arrayOfNodesToConsider.append([consideredRequest,
                                                           allRequests[consideredRequest - 1].deliveryNode])
                            arrayOfNodesToConsider.remove([consideredRequest, consideredNode])
                            currentRoute.append(consideredNode)  # add the node to the path
                            currentRequest.append(consideredRequest)  # add the corresponding request to the path
                            totalLoad += allRequests[consideredRequest - 1].load  # add the corresponding load
                            # add the corresponding time to travel from last node visited to considered Node
                            totalTime += travelTime
                            # add the time wasted waiting for the time windows (applicable only if arrived early)
                            totalTime += wastedTime

                            # we are not considering wasted time as a factor increasing the cost

                            totalCost += 0.0016 * travelTime
                            totalCostEach += 0.0016 * travelTime
                            currentLoad.append(totalLoad)  # adding the corresponding load to the path
                            currentTimeWasted.append(int(wastedTime))  # adding the corresponding waiting time to path
                            currentRouteTime.append(totalTime)

                            return currentRoute, currentRequest, currentLoad, currentTimeWasted, currentRouteTime,\
                                   totalTime, totalLoad, totalRewardEach, totalCostEach, totalReward, totalCost,\
                                   arrayOfNodesToConsider, requestsServed, requestsServedAll, partialServedRequests,\
                                   True

            # if node inside considered request is the delivery node, it has not been served yet, but the corresponding
            # pickup node has been visited
            elif allRequests[consideredRequest - 1].deliveryNode == consideredNode and \
                    consideredRequest not in requestsServed and \
                    consideredRequest in partialServedRequests:
                # check that time windows are not violated
                # capacity constraints cannot be violated since we're dropping off people here, not picking them up
                [checkedTimeWindows, wastedTime, travelTime] = checkTimeWindowsConstraints(consideredRequest,
                                                                               consideredNode, currentRoute, tMatrix,
                                                                               allRequests, totalTime)
                if checkedTimeWindows:  # if time windows are satisfied
                    # try to find a permutation that is feasible for all partially served requests
                    if checkIfRouteIsPossible(currentRoute, consideredRequest, consideredNode, totalTime, travelTime,
                                              wastedTime, timeLimit, tMatrix, driverEnd, allRequests,
                                              partialServedRequests, driverNumber):
                        # remove the request from partially served and add it to served list
                        partialServedRequests.remove(consideredRequest)
                        partialServedRequestsAll.remove(consideredRequest)
                        requestsServed.append(consideredRequest)
                        requestsServedAll.append(consideredRequest)
                        # remove the delivery node from the nodes to consider
                        arrayOfNodesToConsider.remove([consideredRequest, consideredNode])
                        currentRoute.append(consideredNode)  # add the considered node to the route
                        currentRequest.append(consideredRequest)  # adding the corresponding request to the path
                        totalLoad -= allRequests[consideredRequest - 1].load  # update the load
                        totalReward += allRequests[consideredRequest - 1].reward  # update the reward
                        totalRewardEach += allRequests[consideredRequest - 1].reward  # update the reward
                        # add the time required to travel from the last visited node to the considered Node
                        totalTime += travelTime
                        # add the time wasted waiting for time windows (applicable if vehicle arrived early)
                        totalTime += wastedTime
                        # we are not considering wasted time as a cost
                        # we are considering a cost of 10c per minute
                        totalCost += 0.0016 * travelTime
                        totalCostEach += 0.0016 * travelTime
                        currentLoad.append(totalLoad)  # adding the corresponding load to the path
                        currentTimeWasted.append(int(wastedTime))  # adding the corresponding waiting time to the path
                        currentRouteTime.append(totalTime)

                        return currentRoute, currentRequest, currentLoad, currentTimeWasted, currentRouteTime,\
                               totalTime, totalLoad, totalRewardEach, totalCostEach, totalReward, totalCost,\
                               arrayOfNodesToConsider, requestsServed, requestsServedAll, partialServedRequests, True

    return currentRoute, currentRequest, currentLoad, currentTimeWasted, currentRouteTime, totalTime, totalLoad,\
           totalRewardEach, totalCostEach, totalReward, totalCost, arrayOfNodesToConsider, requestsServed,\
           requestsServedAll, partialServedRequests, False  # another node to add was not found


def initialSolution(tMatrix, allRequests, driverNumber, builtRoute, currentRequest, currentLoad, currentTimeWasted,
                    currentRouteTime, allowedCapacity, driverEnd, partialServedRequests, partialServedRequestsAll,
                    dataOfImages):

    # ----------------------------------------- INITIALIZATION OF THE ROUTES -----------------------------------------
    totalTimeEach = []
    timeLimitEach = []
    totalLoadEach = []
    totalRewardEach = []
    totalCostEach = []
    requestsServedEach = []
    requestsServedAll = []

    for j in range(len(driverNumber)):  # for each driver
        # time here refers to the time that the driver starts
        totalTimeEach.append(allRequests[driverNumber[j] - 1].pickupTimeWindowE)
        # the time by which the driver must be at his location
        timeLimitEach.append(allRequests[driverNumber[j] - 1].deliveryTimeWindowL)
        totalLoadEach.append(0)  # load throughout the route
        totalRewardEach.append(0)  # reward for each route
        totalCostEach.append(0)  # cost of each route
        requestsServedEach.append([])

    totalReward = 0  # rewards from requests served
    totalCost = 0  # this is the cost (in cents)
    # requestServed is used to indicate that both the pickup and delivery locations of the request are visited
    requestsServed = []
    # ----------------------------------------------------------------------------------------------------------------

    # CHECK WHICH REQUESTS ARE UNSERVED
    # while continueAddingNodes is true, we continue trying to add new nodes to the route
    continueAddingNodes = [True] * len(driverNumber)
    correspondingRequestNo = [allRequests.requestNo for allRequests in allRequests]
    nodesToConsider = [allRequests.pickupNode for allRequests in allRequests]
    arrayOfNodes = list(zip(correspondingRequestNo, nodesToConsider))
    arrayOfNodesToConsider = []
    for i in range(len(arrayOfNodes)):  # converting tuple to list
        arrayOfNodesToConsider.append(list(arrayOfNodes[i]))

    # if all are false, that means that all nodes were tested for feasibility in all routes and all failed
    while any(x is True for x in continueAddingNodes):
        for j in range(len(driverNumber)):
            if continueAddingNodes[j] is True:
                # -1 since matrix starts from index 0 and the requests start from 1
                # sort the times in ascending order and choose the one with the smallest time cost to be tested for
                # addition to route OR purely randomised by uniform distribution
                listOfNodesRandomised = list(range(0, noOfNodes))
                random.shuffle(listOfNodesRandomised)
                listOfNodesSortedByCost = sorted(range(len(tMatrix[builtRoute[j][-1] - 1])),
                                                 key=lambda k: tMatrix[builtRoute[j][-1] - 1][k])

                # try to find a node to add to the route
                [builtRoute[j], currentRequest[j], currentLoad[j], currentTimeWasted[j], currentRouteTime[j],
                 totalTimeEach[j], totalLoadEach[j], totalRewardEach[j], totalCostEach[j], totalReward, totalCost,
                 arrayOfNodesToConsider, requestsServedEach[j], requestsServedAll, partialServedRequests[j],
                 continueAddingNodes[j]] = \
                    findANode(totalTimeEach[j], totalLoadEach[j], totalRewardEach[j], totalCostEach[j], builtRoute[j],
                              currentRequest[j], currentLoad[j], currentTimeWasted[j], currentRouteTime[j],
                              allowedCapacity, tMatrix, timeLimitEach[j], driverEnd[j], requestsServedEach[j],
                              requestsServedAll, partialServedRequests[j], partialServedRequestsAll,
                              arrayOfNodesToConsider, allRequests, driverNumber[j], listOfNodesRandomised,
                              totalReward, totalCost)

    for j in range(len(driverNumber)):  # for each driver, add their corresponding destination
        travelTimeToDest = tMatrix[builtRoute[j][-1] - 1][allRequests[driverNumber[j] - 1].deliveryNode - 1]
        totalTimeEach[j] += travelTimeToDest  # we update the current time after visiting final node
        # finally, we append the driver's destination node to the route
        builtRoute[j].append(allRequests[driverNumber[j] - 1].deliveryNode)
        totalCostEach[j] += travelTimeToDest * 0.0016
        totalCost += travelTimeToDest * 0.0016

        partialServedRequests[j].remove(driverNumber[j])
        currentRequest[j].append(driverNumber[j])
        currentLoad[j].append(0)
        currentTimeWasted[j].append(0)
        currentRouteTime[j].append(totalTimeEach[j])

    dataOfImages = newSolutionPlot(builtRoute, currentRequest, currentRouteTime, currentLoad, totalReward, totalCost,
                                   dataOfImages)

    return builtRoute, currentRequest, currentLoad, currentTimeWasted, currentRouteTime, totalTimeEach, totalLoadEach,\
           totalCostEach, totalRewardEach, partialServedRequests, requestsServed, requestsServedAll, totalReward,\
           totalCost, dataOfImages


# ------------------------------------------------------- BEGIN -------------------------------------------------------
# initialize the empty arrays to be used to obtain data from the excel sheets
imgNums = np.empty((100, 2))
normGeo = np.empty((100, 2))
distanceMatrix = np.empty((100, 100))
timeMatrix = np.empty((100, 100))
realCoordinates = np.empty((100, 2))

# load the data from the excel
book = load_workbook('information.xlsx')
sheetCoor = book["Coordinates"]
sheetDist = book["distanceMatrix"]
sheetTime = book["distanceTimeMatrix"]

noOfNodes = 100
noOfRequests = 201  # 200 requests
requests = []
data_images = []
maxCPUtime = 120  # 120 seconds

startOfNI = []
startOfGS = []
startOfLS = []
endOfLS = []
bestProfitPlot = []
xProfitPlot = []
percentImprovement = True
driverNo = [1]  # randomly chosen between 1 and 200
capacity = 4
noOfGS = 0

# numsGeo is a reference to the realCoordinates
numsGeo = realCoordinates[:, ::-1]  # reverse the order of the geographical coordinates (from y,x format to x,y format)
# plot the image of Malta to be considered
img = mpimg.imread('maltaMap.PNG')
imgplot = plt.imshow(img)

[imgLimits, geoLimits, geoLeft, geoTop] = obtainScales(plt)  # obtain the scales required from the plot

for i in range(noOfNodes):
    realCoordinates[i, 0] = sheetCoor.cell(row=i + 2, column=1).value
    realCoordinates[i, 1] = sheetCoor.cell(row=i + 2, column=2).value
    # obtain each coordinate, normalise it, convert it to pixels, and plot it on the image
    normGeo[i, :] = normaliseGeo(numsGeo[i, :], geoLeft, geoTop)
    imgNums[i, :] = convertGeoToImg(normGeo[i, :], geoLimits, imgLimits)
    plt.scatter(imgNums[i][0], imgNums[i][1], c='g', s=40)
    for j in range(noOfNodes):  # to populate the distance and time matrices
        distanceMatrix[i, j] = sheetDist.cell(row=i+2, column=j+2).value
        timeMatrix[i, j] = sheetTime.cell(row=i+2, column=j+2).value


# ---------------------------------------------------------------------------------------------------------------------

# greedy search (shortest distance)
print(datetime.datetime.now())
start_time = time.clock()

iterations = 200
kmax = 8
percentNumber = 0.5
monteCarloBook = load_workbook('monteCarloSVinstances.xlsx')
monteCarloResultsBook = load_workbook('monteCarloSVresults.xlsx')
resultsSheet = monteCarloResultsBook['SV Results']

resultsSheet.cell(row=1, column=1).value = 'Iteration No.'
resultsSheet.cell(row=1, column=2).value = 'Route'
resultsSheet.cell(row=1, column=3).value = 'Route Request'
resultsSheet.cell(row=1, column=4).value = 'Profit'
resultsSheet.cell(row=1, column=5).value = 'No. of GS'
resultsSheet.cell(row=1, column=6).value = 'Start Time'
resultsSheet.cell(row=1, column=7).value = 'End Time'
resultsSheet.cell(row=1, column=8).value = 'Execution Time'
resultsSheet.cell(row=1, column=9).value = 'Time to find Best Solution'
resultsSheet.cell(row=1, column=10).value = 'Driver Time (in minutes)'

startingTime = 0
eMinTimeDeviation = 0  # 0 minutes
eMaxTimeDeviation = 7200  # 120 minutes
lMinTimeDeviation = 600  # 10 minutes
lMaxTimeDeviation = 3600  # 60 minutes
endingTime = 9000  # 150 minutes

for i in range(iterations):

    currentSheet = monteCarloBook.create_sheet('Test Instance ' + str(i+1))

    loadOfRequests = np.random.poisson(1, (noOfRequests, 1))  # poisson distribution for the load with expected value 1

    currentSheet.cell(row=1, column=1).value = 'Request No.'
    currentSheet.cell(row=1, column=2).value = 'Origin'
    currentSheet.cell(row=1, column=3).value = 'Destination'
    currentSheet.cell(row=1, column=4).value = 'Earliest Time (at Pickup)'
    currentSheet.cell(row=1, column=5).value = 'Latest Time (at Pickup)'
    currentSheet.cell(row=1, column=6).value = 'Earliest Time (at Destination)'
    currentSheet.cell(row=1, column=7).value = 'Latest Time (at Destination)'
    currentSheet.cell(row=1, column=8).value = 'Distance to Destination (in time)'
    currentSheet.cell(row=1, column=9).value = 'Cost of Request'
    currentSheet.cell(row=1, column=10).value = 'Load'

    requests = []

    driverTime = random.randint(3600, 7200)

    for ij in range(noOfRequests):
        requestNo = ij + 1
        # pick any two nodes without replacement
        directTime = 0
        pickupNode = 0
        deliveryNode = 0

        # requests must be at least 8 minutes in length - approx 5km
        while directTime < 480:  # 8 minutes
            [pickupNode, deliveryNode] = np.random.choice(range(1, noOfNodes + 1), 2, replace=False)
            directTime = timeMatrix[pickupNode - 1, deliveryNode - 1]  # array starts from 0, not from 1

        pickupTimeWindowE = 10000
        while pickupTimeWindowE + directTime > 9000:
            pickupTimeWindowE = startingTime + random.randint(eMinTimeDeviation, eMaxTimeDeviation)
        if ij == 0:  # making driver's time as specified by the driverTime parameter
            while pickupTimeWindowE + driverTime > 9000:
                pickupTimeWindowE = startingTime + random.randint(eMinTimeDeviation, eMaxTimeDeviation)
        deliveryTimeWindowE = pickupTimeWindowE + directTime
        deliveryTimeWindowL = pickupTimeWindowE + directTime + random.randint(lMinTimeDeviation, lMaxTimeDeviation)
        if ij == 0:
            deliveryTimeWindowL = pickupTimeWindowE + driverTime
        if deliveryTimeWindowL > 9000:
            deliveryTimeWindowL = 9000

        pickupTimeWindowL = deliveryTimeWindowL - directTime
        directDistance = directTime

        load = loadOfRequests[ij, 0]
        if load < 1:
            load = 1
        elif load > 4:
            load = 4

        costVariance = random.uniform(-0.7,
                                      0.7)  # cost is assumed to vary between -70c and 70c (some people are generous)

        costOfRequest = round((directTime / 60 * 0.15 + costVariance) * 100) / 100  # 15 cents per minute
        if costOfRequest < 1:
            costOfRequest = 1
        elif costOfRequest > 4:
            costOfRequest = 4

        # a reduced multiplier for people travelling together
        if loadOfRequests[ij, 0] == 2:
            costOfRequest = costOfRequest * 1.8
        elif loadOfRequests[ij, 0] == 3:
            costOfRequest = costOfRequest * 2.6
        elif loadOfRequests[ij, 0] == 4:
            costOfRequest = costOfRequest * 3.4

        reward = costOfRequest

        r1 = Request(requestNo, pickupNode, deliveryNode, pickupTimeWindowE, pickupTimeWindowL, deliveryTimeWindowE,
                     deliveryTimeWindowL, directDistance, reward, load)
        requests.append(r1)

        currentSheet.cell(row=ij + 2, column=1).value = requestNo
        currentSheet.cell(row=ij + 2, column=2).value = pickupNode
        currentSheet.cell(row=ij + 2, column=3).value = deliveryNode
        currentSheet.cell(row=ij + 2, column=4).value = pickupTimeWindowE
        currentSheet.cell(row=ij + 2, column=5).value = pickupTimeWindowL
        currentSheet.cell(row=ij + 2, column=6).value = deliveryTimeWindowE
        currentSheet.cell(row=ij + 2, column=7).value = deliveryTimeWindowL
        currentSheet.cell(row=ij + 2, column=8).value = directTime
        currentSheet.cell(row=ij + 2, column=9).value = costOfRequest
        currentSheet.cell(row=ij + 2, column=10).value = load


    print("iteration is:", i)
    timeToFindBestSolution = 0
    route = []
    routeRequest = []
    routeLoad = []
    routeTimeWasted = []
    routeTime = []
    partiallyServedRequests = []
    partiallyServedRequestsAllVehicles = []
    driverEndNode = []

    for ii in range(len(driverNo)):
        route.append([requests[driverNo[ii] - 1].pickupNode])
        routeRequest.append([requests[driverNo[ii] - 1].requestNo])
        routeLoad.append([0])
        routeTimeWasted.append([0])
        routeTime.append([requests[driverNo[ii] - 1].pickupTimeWindowE])
        partiallyServedRequests.append([driverNo[ii]])
        partiallyServedRequestsAllVehicles.append(driverNo[ii])
        driverEndNode.append(requests[driverNo[ii] - 1].deliveryNode)

    [route, routeRequest, routeLoad, routeTimeWasted, routeTime, finalTime, finalLoad, finalCostEach, finalRewardEach,
     partiallyServedRequests, servedRequests, servedRequestsAll, finalReward, finalCost, dataOfImages] = \
        initialSolution(timeMatrix, requests, driverNo, route, routeRequest, routeLoad, routeTimeWasted, routeTime,
                        capacity, driverEndNode, partiallyServedRequests, partiallyServedRequestsAllVehicles,
                        data_images)

    [route, routeRequest, routeLoad, routeTimeWasted, routeTime, routeReward, routeCost, finalReward, finalCost,
     partiallyServedRequests, servedRequests, partiallyServedRequestsAllVehicles, servedRequestsAll,
     isPossible] = \
        updateAllRoutes(route, routeRequest, routeTime, timeMatrix, requests)

    [route, routeRequest, routeLoad, routeTime, servedRequests, servedRequestsAll, routeReward, routeCost,
     dataOfImages] = \
        localSearch(route, routeRequest, routeTime, routeLoad, routeReward, routeCost, servedRequests,
                    servedRequestsAll, timeMatrix,
                    requests, noOfRequests, dataOfImages)

    [route, routeRequest, routeLoad, routeTimeWasted, routeTime, routeReward, routeCost, finalReward, finalCost,
     partiallyServedRequests, servedRequests, partiallyServedRequestsAllVehicles, servedRequestsAll,
     isPossible] = \
        updateAllRoutes(route, routeRequest, routeTime, timeMatrix, requests)

    start_time_VNS = time.clock()
    [route, routeRequest, numGS, timeToFindBestSolution] = VNS(route, routeRequest, routeTime, routeLoad,
                                                               servedRequests, routeReward,
                                                               routeCost,
                                                               timeMatrix, requests, noOfRequests, data_images)

    end_time_VNS = time.clock()
    [route, routeRequest, routeLoad, routeTimeWasted, routeTime, routeReward, routeCost, finalReward, finalCost,
     partiallyServedRequests, servedRequests, partiallyServedRequestsAllVehicles, servedRequestsAll,
     isPossible] = \
        updateAllRoutes(route, routeRequest, routeTime, timeMatrix, requests)

    dataOfImages = newSolutionPlot(route, routeRequest, routeTime, routeLoad, finalReward, finalCost, dataOfImages)

    resultsSheet.cell(row=i + 2, column=1).value = i + 1
    resultsSheet.cell(row=i + 2, column=2).value = str(route)
    resultsSheet.cell(row=i + 2, column=3).value = str(routeRequest)
    resultsSheet.cell(row=i + 2, column=4).value = finalReward - finalCost
    resultsSheet.cell(row=i + 2, column=5).value = numGS
    resultsSheet.cell(row=i + 2, column=6).value = start_time_VNS
    resultsSheet.cell(row=i + 2, column=7).value = end_time_VNS
    resultsSheet.cell(row=i + 2, column=8).value = end_time_VNS - start_time_VNS
    resultsSheet.cell(row=i + 2, column=9).value = timeToFindBestSolution
    resultsSheet.cell(row=i + 2, column=10).value = round(driverTime/60, 1)

    monteCarloBook.save('monteCarloSVinstances.xlsx')
    monteCarloResultsBook.save('monteCarloSVresults.xlsx')

dataOfImages = newSolutionPlot(route, routeRequest, routeTime, routeLoad, finalReward, finalCost, dataOfImages)

print(datetime.datetime.now())
end_time = time.clock()
execTime = end_time - start_time
print("--- %s seconds ---" % (execTime))
