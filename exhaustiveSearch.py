import numpy as np
from openpyxl import Workbook, load_workbook
import itertools
import datetime
import time
import matplotlib.pyplot as plt
import matplotlib.image as mpimg


class Request:
    def __init__(self, requestNo, pickupNode, deliveryNode, pickupTimeWindowE, pickupTimeWindowL, deliveryTimeWindowE,
                 deliveryTimeWindowL, directDistance, reward, load):
        self.requestNo = requestNo
        self.pickupNode = pickupNode
        self.deliveryNode = deliveryNode
        self.pickupTimeWindowE = pickupTimeWindowE
        self.pickupTimeWindowL = pickupTimeWindowL
        self.deliveryTimeWindowE = deliveryTimeWindowE
        self.deliveryTimeWindowL = deliveryTimeWindowL
        self.directDistance = directDistance
        self.reward = reward
        self.load = load

class unique_element:
    def __init__(self,value,occurrences):
        self.value = value
        self.occurrences = occurrences

def perm_unique(elements):
    eset=set(elements)
    listunique = [unique_element(i,elements.count(i)) for i in eset]
    u=len(elements)
    return perm_unique_helper(listunique,[0]*u,u-1)

def perm_unique_helper(listunique,result_list,d):
    if d < 0:
        yield tuple(result_list)
    else:
        for i in listunique:
            if i.occurrences > 0:
                result_list[d]=i.value
                i.occurrences-=1
                for g in  perm_unique_helper(listunique,result_list,d-1):
                    yield g
                i.occurrences+=1


def obtainScales (pltMalta):
    # taking the size of the image
    bottom, top = pltMalta.ylim()
    left, right = pltMalta.xlim()

    bottom += 0.5
    right += 0.5

    # x, y format
    imageLimits = np.array([right, bottom])  # scale of image coordinates

    # fixed coordinates of the image shown
    geographicBottom = 35.848350
    geographicTop = 35.932180
    geographicLeft = 14.399522
    geographicRight = 14.539597

    # x, y format
    # scale of geographic coordinates
    geographicLimits = np.array([geographicRight - geographicLeft, geographicBottom - geographicTop])
    return imageLimits, geographicLimits, geographicLeft, geographicTop

def normaliseGeo (geoLocation, geoL, geoT):
    # used for finding a linear relationship between geographical coordinates and image pixels
    location = np.array([geoLocation[0] - geoL, geoLocation[1] - geoT])
    return location

def convertGeoToImg (geoLocation, geoScale, imgScale):
    # find the image pixels of where the geographical coordinates are pointing to
    imgLocation = np.multiply(np.divide(geoLocation, geoScale), imgScale)
    return imgLocation

def convertImgToGeo (imgLocation, geoScale, imgScale, geoT, geoL):
    # find the geographical coordinates of an image pixel to find its respective point on map
    geoLocation = np.multiply(np.divide(imgLocation, imgScale), geoScale) + np.array([geoL, geoT])
    return geoLocation


def checkCapacityConstraints(consideredRequest, allowedCapacity, totalLoad, allRequests):
    # we want to make sure that adding the load of the considered node to the vehicle will not exceed capacity
    # load total = load before arriving at node
    # allRequests...load = load at considered Node
    return totalLoad + allRequests[consideredRequest - 1].load <= allowedCapacity

def checkTimeWindowsConstraints(consideredRequest, consideredNode, currentRoute, tMatrix, allRequests, currentTime):
    checked = True  # if this is false, that means that the node considered could not be added due to time windows
    currTime = currentTime  # this is used as a dummy variable for the total Time of the route
    timeWasted = 0  # this is used as an indication of the total time left waiting for a customer in a pickup location
    e1 = 0
    l1 = 0
    travelTime = tMatrix[currentRoute[-1] - 1][consideredNode - 1]
    currTime += travelTime
    # time windows of the next node to be added to the route
    if consideredNode == allRequests[consideredRequest - 1].pickupNode:
        e1 = allRequests[consideredRequest - 1].pickupTimeWindowE
        l1 = allRequests[consideredRequest - 1].pickupTimeWindowL
    elif consideredNode == allRequests[consideredRequest - 1].deliveryNode:
        e1 = allRequests[consideredRequest - 1].deliveryTimeWindowE
        l1 = allRequests[consideredRequest - 1].deliveryTimeWindowL
    if currTime < e1:  # it must wait and time is wasted
        timeWasted += e1 - currTime
        checked = True
    elif currTime <= l1:  # on time
        checked = True
    else:  # if time window not reached
        checked = False

    return checked, timeWasted, travelTime


def updateRouteInfo(currentRoute, currentRequest, currentRouteTime, tMatrix, allRequests):
    # USED TO UPDATE THE INFORMATION ON A SINGLE ROUTE
    # ------------------------------------------------ INITIALIZATION ------------------------------------------------
    tempRoute = []
    tempRouteRequest = []
    tempRouteLoad = []
    tempRouteTimeWasted = []
    tempRouteTime = []
    tempServedRequests = []
    tempPartiallyServedRequests = []
    tempFinalReward = 0
    tempFinalCost = 0

    tempRoute.append(currentRoute[0])
    tempRouteRequest.append(currentRequest[0])
    tempRouteLoad.append(0)
    tempRouteTimeWasted.append(0)
    tempRouteTime.append(currentRouteTime[0])
    tempPartiallyServedRequests.append(currentRequest[0])
    # ----------------------------------------------------------------------------------------------------------------

    for i in range(1, len(currentRoute)):  # testing all the nodes inside the route
        # CAPACITY
        if allRequests[currentRequest[i] - 1].pickupNode == currentRoute[i]:  # if it is a pickup node
            tempRouteLoad.append(tempRouteLoad[-1] + allRequests[currentRequest[i] - 1].load)  # increase the load
            tempPartiallyServedRequests.append(currentRequest[i])
            if tempRouteLoad[-1] > capacity:  # if capacity constraints have been violated
                # exit the function with isItPossible = False
                return currentRoute, currentRequest, tempRouteLoad, tempRouteTimeWasted, tempRouteTime,\
                       tempFinalReward, tempFinalCost, tempPartiallyServedRequests, tempServedRequests, False

        # capacity constraints cannot be violated if we are at a delivery node
        elif allRequests[currentRequest[i] - 1].deliveryNode == currentRoute[i]:
            tempServedRequests.append(currentRequest[i])
            tempPartiallyServedRequests.remove(currentRequest[i])

            # if request considered is the driver's request, then no reward is given
            if currentRequest[i] == currentRequest[0]:
                tempRouteLoad.append(0)
            else:  # update the load and reward
                tempRouteLoad.append(tempRouteLoad[-1] - allRequests[currentRequest[i] - 1].load)
                tempFinalReward += allRequests[currentRequest[i] - 1].reward  # update the reward

        # TIME WINDOWS
        [timeWindowsCheck, temporaryTimeWasted, timeToTravel] =\
                                    checkTimeWindowsConstraints(currentRequest[i], currentRoute[i], tempRoute, tMatrix,
                                                                allRequests, tempRouteTime[-1])
        tempRouteTimeWasted.append(temporaryTimeWasted)  # update time wasted list

        if timeWindowsCheck:  # if time windows constraints are not violated
            tempRouteTime.append(tempRouteTime[-1] + tempRouteTimeWasted[i] + timeToTravel)  # update time list
        else:  # time windows constraints have been violated
            # exit the function with isItPossible = False
            return currentRoute, currentRequest, tempRouteLoad, tempRouteTimeWasted, tempRouteTime, tempFinalReward,\
                   tempFinalCost, tempPartiallyServedRequests, tempServedRequests, False

        # we are not considering wasted time as a cost
        # cost is calculated to be 10c per minute
        tempFinalCost += 0.0016 * timeToTravel
        tempRoute.append(currentRoute[i])
        tempRouteRequest.append(currentRequest[i])

    return tempRoute, tempRouteRequest, tempRouteLoad, tempRouteTimeWasted, tempRouteTime, tempFinalReward,\
           tempFinalCost, tempPartiallyServedRequests, tempServedRequests, True


# ------------------------------------------------------- BEGIN ------------------------------------------------------ #
# initialize the empty arrays to be used to obtain data from the excel sheets
imgNums = np.empty((100, 2))
normGeo = np.empty((100, 2))
distanceMatrix = np.empty((100, 100))
timeMatrix = np.empty((100, 100))
realCoordinates = np.empty((100, 2))

# load the data from the excel
book = load_workbook('information.xlsx')
sheetCoor = book["Coordinates"]
sheetDist = book["distanceMatrix"]
sheetTime = book["distanceTimeMatrix"]

instanceNo = 0
exhaustiveSearchInstances = load_workbook('exhaustiveSearchInstances.xlsx')
sheetRequests = exhaustiveSearchInstances["Test Instance " + str(instanceNo + 1)]

noOfNodes = 100
noOfRequests = 41  # 40 requests
requests = []
data_images = []

driverNo = 1  # randomly chosen between 1 and 200
capacity = 4

# numsGeo is a reference to the realCoordinates
numsGeo = realCoordinates[:, ::-1]  # reverse the order of the geographical coordinates (from y,x format to x,y format)
# plot the image of Malta to be considered
img = mpimg.imread('maltaMap.PNG')
imgplot = plt.imshow(img)

[imgLimits, geoLimits, geoLeft, geoTop] = obtainScales(plt)  # obtain the scales required from the plot

for i in range(noOfNodes):
    realCoordinates[i, 0] = sheetCoor.cell(row=i + 2, column=1).value
    realCoordinates[i, 1] = sheetCoor.cell(row=i + 2, column=2).value
    # obtain each coordinate, normalise it, convert it to pixels, and plot it on the image
    normGeo[i, :] = normaliseGeo(numsGeo[i, :], geoLeft, geoTop)
    imgNums[i, :] = convertGeoToImg(normGeo[i, :], geoLimits, imgLimits)
    plt.scatter(imgNums[i][0], imgNums[i][1], c='g', s=40)
    for j in range(noOfNodes):  # to populate the distance and time matrices
        distanceMatrix[i, j] = sheetDist.cell(row=i+2, column=j+2).value
        timeMatrix[i, j] = sheetTime.cell(row=i+2, column=j+2).value

for i in range(noOfRequests):
    requestNo = sheetRequests.cell(row=i+2, column=1).value
    pickupNode = sheetRequests.cell(row=i+2, column=2).value
    deliveryNode = sheetRequests.cell(row=i+2, column=3).value
    pickupTimeWindowE = sheetRequests.cell(row=i+2, column=4).value
    pickupTimeWindowL = sheetRequests.cell(row=i+2, column=5).value
    deliveryTimeWindowE = sheetRequests.cell(row=i+2, column=6).value
    deliveryTimeWindowL = sheetRequests.cell(row=i+2, column=7).value
    directDistance = sheetRequests.cell(row=i+2, column=8).value
    reward = sheetRequests.cell(row=i+2, column=9).value
    load = sheetRequests.cell(row=i+2, column=10).value

    r1 = Request(requestNo, pickupNode, deliveryNode, pickupTimeWindowE, pickupTimeWindowL, deliveryTimeWindowE,
                 deliveryTimeWindowL, directDistance, reward, load)
    requests.append(r1)

route = []
route.append(requests[driverNo - 1].pickupNode)
# route.append(requests[driverNo - 1].deliveryNode)
driverEndNode = requests[driverNo - 1].deliveryNode

routeRequest = []
routeRequest.append(requests[driverNo - 1].requestNo)

routeLoad = []
routeLoad.append(0)

routeTimeWasted = []
routeTimeWasted.append(0)

routeTime = []
routeTime.append(requests[driverNo - 1].pickupTimeWindowE)

timeToFindBestProfit = 0
requestNumbers = list(range(1, noOfRequests + 1))
requestNumbers.remove(requests[driverNo - 1].requestNo)

print("start time")
print(datetime.datetime.now())
start_time = time.clock()
bestRoute = [requests[driverNo - 1].pickupNode, requests[driverNo - 1].deliveryNode]
bestRouteRequest = [requests[driverNo - 1].requestNo, requests[driverNo - 1].requestNo]
bestReward = 0
bestCost = timeMatrix[bestRoute[0] - 1][bestRoute[1] - 1] * 0.0016
continueLoop = True

# Exhaustive Search
for i in range(noOfRequests):

    if not continueLoop:
        break

    print("considering ", i+1, " requests")
    route = []
    route.append(requests[driverNo - 1].pickupNode)

    routeRequest = []
    routeRequest.append(requests[driverNo - 1].requestNo)

    routeLoad = []
    routeLoad.append(0)

    routeTimeWasted = []
    routeTimeWasted.append(0)

    routeTime = []
    routeTime.append(requests[driverNo - 1].pickupTimeWindowE)

    exhaustiveList = list(itertools.combinations(requestNumbers, r=i+1))
    listToConsider = []

    continueLoop = False

    for j in range(len(exhaustiveList)):

        listToConsider.append(list(exhaustiveList[j]))
        for k in range(i+1):
            listToConsider[j].append(exhaustiveList[j][k])

        availablePermut = list(perm_unique(listToConsider[j]))

        for m in range(len(availablePermut)):

            # route construction
            routeRequest = []
            routeRequest.append(requests[driverNo - 1].requestNo)
            for n in range(len(list(availablePermut[m]))):
                routeRequest.append(availablePermut[m][n])
            routeRequest.append(requests[driverNo - 1].requestNo)

            route = []
            partiallySerReq = []
            for p in range(len(routeRequest)):
                if routeRequest[p] not in partiallySerReq:
                    route.append(requests[routeRequest[p] - 1].pickupNode)
                    partiallySerReq.append(routeRequest[p])
                else:
                    route.append(requests[routeRequest[p] - 1].deliveryNode)

            # check profit and feasibility of route
            [route, routeRequest, routeLoad, routeTimeWasted, routeTime, finalReward, finalCost,
             partiallyServedRequests, servedRequests, isPossible] = \
                updateRouteInfo(route, routeRequest, routeTime, timeMatrix, requests)

            # if a feasible route is found
            if isPossible:
                continueLoop = True
                if finalReward - finalCost > bestReward - bestCost:
                    bestRoute = route[:]
                    bestRouteRequest = routeRequest[:]
                    bestReward = finalReward
                    bestCost = finalCost
                    print("best profit so far:", finalReward - finalCost)
                    timeToFindBestProfit = time.clock() - start_time

print(" ")
print("best route:        ", bestRoute)
print("best route request:", bestRouteRequest)
print("best Profit:       ", bestReward - bestCost)
print(" ")
print(datetime.datetime.now())
end_time = time.clock()
execTime = end_time - start_time
print("--- %s seconds ---" % (execTime))
bookResults = load_workbook('VNSvsExhaustiveSearch.xlsx')
sheetES = bookResults['Exhaustive Search Results']
sheetES.cell(row=instanceNo + 2, column=1).value = instanceNo
sheetES.cell(row=instanceNo + 2, column=2).value = str(bestRoute)
sheetES.cell(row=instanceNo + 2, column=3).value = str(bestRouteRequest)
sheetES.cell(row=instanceNo + 2, column=4).value = bestReward - bestCost
sheetES.cell(row=instanceNo + 2, column=6).value = start_time
sheetES.cell(row=instanceNo + 2, column=7).value = end_time
sheetES.cell(row=instanceNo + 2, column=8).value = end_time - start_time
sheetES.cell(row=instanceNo + 2, column=9).value = timeToFindBestProfit
bookResults.save('VNSvsExhaustiveSearch.xlsx')
